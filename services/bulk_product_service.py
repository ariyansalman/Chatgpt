"""Bulk Product Import & Export service — V35.

Handles CSV, Excel (.xlsx), and JSON import/export for products.
Provides bulk action execution with validation and audit logging.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ─── Column mapping for import/export ─────────────────────────────────────

PRODUCT_EXPORT_FIELDS = [
    "id", "name", "description", "category", "subcategory",
    "price", "sale_price", "currency", "product_type",
    "stock_count", "is_active", "is_featured",
    "delivery_note", "warranty_info",
    "min_quantity", "max_quantity",
    "bulk_purchase_enabled", "reusable",
    "product_emoji", "sort_order", "created_at",
]

PRODUCT_IMPORT_REQUIRED = {"name", "price", "product_type"}

VALID_PRODUCT_TYPES = {
    "KEY", "REDEEM_LINK", "ACCOUNT_LOGIN", "VOUCHER",
    "FILE", "DOWNLOADABLE_FILE", "AUTO_GENERATED",
    "MANUAL_DELIVERY", "PREORDER", "SUBSCRIPTION",
    "BUNDLE", "SERVICE", "EXTERNAL_DELIVERY",
}


# ─── Import helpers ────────────────────────────────────────────────────────

def _parse_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "on")


def _parse_float(v: Any) -> Optional[float]:
    try:
        return float(str(v).strip()) if v not in (None, "", "null") else None
    except (ValueError, TypeError):
        return None


def _parse_int(v: Any) -> Optional[int]:
    try:
        return int(str(v).strip()) if v not in (None, "", "null") else None
    except (ValueError, TypeError):
        return None


def _validate_row(row: Dict[str, Any], row_num: int) -> Tuple[bool, str]:
    """Returns (is_valid, error_message)."""
    missing = PRODUCT_IMPORT_REQUIRED - set(k for k, v in row.items() if v not in (None, ""))
    if missing:
        return False, f"Row {row_num}: missing required fields: {', '.join(missing)}"

    price = _parse_float(row.get("price"))
    if price is None or price < 0:
        return False, f"Row {row_num}: invalid price '{row.get('price')}'"

    ptype = str(row.get("product_type", "")).strip().upper()
    if ptype not in VALID_PRODUCT_TYPES:
        return False, f"Row {row_num}: invalid product_type '{ptype}'"

    sale_price = _parse_float(row.get("sale_price"))
    if sale_price is not None and sale_price < 0:
        return False, f"Row {row_num}: sale_price must be >= 0"

    stock = _parse_int(row.get("stock_count"))
    if stock is not None and stock < 0:
        return False, f"Row {row_num}: stock_count must be >= 0"

    return True, ""


# ─── Format parsers ────────────────────────────────────────────────────────

def parse_csv_products(file_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse CSV bytes → list of product dicts + list of parse errors."""
    rows: List[Dict] = []
    errors: List[str] = []
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            rows.append({k.strip().lower(): v for k, v in row.items()})
    except Exception as e:
        errors.append(f"CSV parse error: {e}")
    return rows, errors


def parse_json_products(file_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse JSON bytes → list of product dicts + list of parse errors."""
    rows: List[Dict] = []
    errors: List[str] = []
    try:
        data = json.loads(file_bytes.decode("utf-8", errors="replace"))
        if isinstance(data, list):
            rows = [{k.strip().lower(): v for k, v in r.items()} for r in data if isinstance(r, dict)]
        elif isinstance(data, dict) and "products" in data:
            rows = [{k.strip().lower(): v for k, v in r.items()} for r in data["products"] if isinstance(r, dict)]
        else:
            errors.append("JSON must be a list of products or {'products': [...]}")
    except json.JSONDecodeError as e:
        errors.append(f"JSON parse error: {e}")
    return rows, errors


def parse_xlsx_products(file_bytes: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse Excel bytes → list of product dicts + list of parse errors."""
    rows: List[Dict] = []
    errors: List[str] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for i, sheet_row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower() if h is not None else "" for h in sheet_row]
                continue
            row_dict = dict(zip(headers, sheet_row))
            rows.append(row_dict)
        wb.close()
    except ImportError:
        errors.append("Excel support requires openpyxl — install it with: pip install openpyxl")
    except Exception as e:
        errors.append(f"Excel parse error: {e}")
    return rows, errors


def parse_product_file(file_bytes: bytes, file_format: str) -> Tuple[List[Dict], List[str]]:
    """Dispatch to the right parser based on format."""
    fmt = file_format.lower().strip(".")
    if fmt == "csv":
        return parse_csv_products(file_bytes)
    if fmt in ("xlsx", "xls"):
        return parse_xlsx_products(file_bytes)
    if fmt == "json":
        return parse_json_products(file_bytes)
    return [], [f"Unsupported format: {file_format}. Use csv, xlsx, or json."]


# ─── Category resolution ───────────────────────────────────────────────────

def _resolve_category(session, name: str) -> Optional[int]:
    from database.models import Category
    if not name:
        return None
    cat = session.query(Category).filter(Category.name.ilike(name.strip())).first()
    return cat.id if cat else None


def _resolve_subcategory(session, name: str, cat_id: Optional[int]) -> Optional[int]:
    from database.models import Subcategory
    if not name:
        return None
    q = session.query(Subcategory).filter(Subcategory.name.ilike(name.strip()))
    if cat_id:
        q = q.filter(Subcategory.category_id == cat_id)
    sub = q.first()
    return sub.id if sub else None


# ─── Product duplicate detection ───────────────────────────────────────────

def _find_duplicate_product(session, name: str, cat_id: Optional[int]) -> bool:
    from database.models import Product
    q = session.query(Product).filter(Product.name.ilike(name.strip()))
    if cat_id:
        q = q.filter(Product.category_id == cat_id)
    return q.first() is not None


# ─── Import execution ──────────────────────────────────────────────────────

def import_products(
    file_bytes: bytes,
    file_format: str,
    admin_id: int,
    max_rows: int = 1000,
) -> Dict[str, Any]:
    """
    Parse and import products from uploaded file.

    Returns a report dict:
      total_rows, imported, failed, duplicates, errors, record_id
    """
    from database import get_db_session
    from database.models import Product, ProductType, BulkImportRecord

    report: Dict[str, Any] = {
        "total_rows": 0, "imported": 0, "failed": 0,
        "duplicates": 0, "errors": [],
    }

    # Create tracking record
    record_id: Optional[int] = None
    try:
        with get_db_session() as s:
            rec = BulkImportRecord(
                admin_id=admin_id,
                file_format=file_format.lower(),
                status="RUNNING",
                started_at=datetime.utcnow(),
            )
            s.add(rec)
            s.flush()
            record_id = rec.id
    except Exception as e:
        logger.exception("Could not create BulkImportRecord")

    rows, parse_errors = parse_product_file(file_bytes, file_format)
    report["errors"].extend(parse_errors)

    if not rows and parse_errors:
        _finish_import_record(record_id, report, "FAILED")
        return report

    # Enforce row limit
    if len(rows) > max_rows:
        report["errors"].append(
            f"File has {len(rows)} rows but limit is {max_rows}. Truncating."
        )
        rows = rows[:max_rows]

    report["total_rows"] = len(rows)

    with get_db_session() as session:
        for i, row in enumerate(rows, start=1):
            valid, err = _validate_row(row, i)
            if not valid:
                report["failed"] += 1
                report["errors"].append(err)
                continue

            name = str(row.get("name", "")).strip()
            cat_name = str(row.get("category", "")).strip()
            sub_name = str(row.get("subcategory", "")).strip()

            cat_id = _resolve_category(session, cat_name) if cat_name else None
            sub_id = _resolve_subcategory(session, sub_name, cat_id) if sub_name else None

            # Duplicate check
            if _find_duplicate_product(session, name, cat_id):
                report["duplicates"] += 1
                report["errors"].append(f"Row {i}: duplicate product name '{name}' — skipped")
                continue

            ptype_str = str(row.get("product_type", "KEY")).strip().upper()
            try:
                ptype = ProductType[ptype_str]
            except KeyError:
                report["failed"] += 1
                report["errors"].append(f"Row {i}: unknown product_type '{ptype_str}'")
                continue

            price = _parse_float(row.get("price")) or 0.0
            sale_price = _parse_float(row.get("sale_price"))
            stock = _parse_int(row.get("stock_count")) or 0
            currency = str(row.get("currency", "USD")).strip().upper() or "USD"
            is_active = _parse_bool(row.get("is_active", True))
            is_featured = _parse_bool(row.get("is_featured", False))
            delivery_note = str(row.get("delivery_note", "")).strip() or None
            warranty_info = str(row.get("warranty_info", "")).strip() or None
            min_qty = _parse_int(row.get("min_quantity"))
            max_qty = _parse_int(row.get("max_quantity"))
            bulk_enabled = _parse_bool(row.get("bulk_purchase_enabled", True))
            reusable = _parse_bool(row.get("reusable", False))
            emoji = str(row.get("product_emoji", "")).strip() or None
            sort_order = _parse_int(row.get("sort_order"))
            desc = str(row.get("description", "")).strip() or None

            product = Product(
                name=name,
                description=desc,
                price=price,
                sale_price=sale_price,
                stock_count=stock,
                product_type=ptype,
                category_id=cat_id,
                subcategory_id=sub_id,
                currency=currency,
                is_active=is_active,
                is_featured=is_featured,
                delivery_note=delivery_note,
                warranty_info=warranty_info,
                min_quantity=min_qty,
                max_quantity=max_qty,
                bulk_purchase_enabled=bulk_enabled,
                reusable=reusable,
                product_emoji=emoji,
                sort_order=sort_order,
                created_at=datetime.utcnow(),
            )
            try:
                session.add(product)
                session.flush()
                report["imported"] += 1
            except Exception as e:
                session.rollback()
                report["failed"] += 1
                report["errors"].append(f"Row {i}: DB error — {e}")

    _finish_import_record(record_id, report, "COMPLETED")
    return report


def _finish_import_record(record_id: Optional[int], report: Dict, status: str) -> None:
    if record_id is None:
        return
    from database import get_db_session
    from database.models import BulkImportRecord
    try:
        with get_db_session() as s:
            rec = s.query(BulkImportRecord).filter_by(id=record_id).first()
            if rec:
                rec.status = status
                rec.total_rows = report.get("total_rows", 0)
                rec.imported = report.get("imported", 0)
                rec.failed = report.get("failed", 0)
                rec.duplicates = report.get("duplicates", 0)
                rec.report = json.dumps(report.get("errors", [])[:100])
                rec.completed_at = datetime.utcnow()
                if report.get("errors"):
                    rec.error_summary = "; ".join(report["errors"][:3])[:500]
    except Exception:
        logger.exception("Could not update BulkImportRecord id=%s", record_id)


# ─── Export builders ───────────────────────────────────────────────────────

def _product_to_dict(product, cat_map: Dict[int, str], sub_map: Dict[int, str]) -> Dict:
    return {
        "id": product.id,
        "name": product.name,
        "description": product.description or "",
        "category": cat_map.get(product.category_id, ""),
        "subcategory": sub_map.get(product.subcategory_id, ""),
        "price": product.price,
        "sale_price": product.sale_price or "",
        "currency": product.currency,
        "product_type": product.product_type.name if product.product_type else "",
        "stock_count": product.stock_count,
        "is_active": product.is_active,
        "is_featured": product.is_featured,
        "delivery_note": product.delivery_note or "",
        "warranty_info": product.warranty_info or "",
        "min_quantity": product.min_quantity or "",
        "max_quantity": product.max_quantity or "",
        "bulk_purchase_enabled": product.bulk_purchase_enabled,
        "reusable": product.reusable,
        "product_emoji": product.product_emoji or "",
        "sort_order": product.sort_order or "",
        "created_at": product.created_at.strftime("%Y-%m-%d %H:%M:%S") if product.created_at else "",
    }


def _fetch_products(scope: str, scope_arg: Optional[str] = None) -> List:
    from database import get_db_session
    from database.models import Product
    with get_db_session() as s:
        q = s.query(Product)
        if scope == "category" and scope_arg:
            try:
                q = q.filter(Product.category_id == int(scope_arg))
            except ValueError:
                pass
        elif scope == "selected" and scope_arg:
            try:
                ids = [int(x) for x in scope_arg.split(",") if x.strip()]
                q = q.filter(Product.id.in_(ids))
            except ValueError:
                pass
        products = q.order_by(Product.id).all()
        # Detach safely
        return [
            {col.name: getattr(p, col.name) for col in p.__table__.columns}
            | {
                "product_type_name": p.product_type.name if p.product_type else "",
                "category_id": p.category_id,
                "subcategory_id": p.subcategory_id,
                "is_active": p.is_active,
                "is_featured": p.is_featured,
                "created_at_str": p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else "",
            }
            for p in products
        ]


def _build_cat_maps() -> Tuple[Dict[int, str], Dict[int, str]]:
    from database import get_db_session
    from database.models import Category, Subcategory
    cat_map: Dict[int, str] = {}
    sub_map: Dict[int, str] = {}
    with get_db_session() as s:
        for c in s.query(Category).all():
            cat_map[c.id] = c.name
        for sc in s.query(Subcategory).all():
            sub_map[sc.id] = sc.name
    return cat_map, sub_map


def export_products_csv(scope: str = "all", scope_arg: Optional[str] = None) -> bytes:
    cat_map, sub_map = _build_cat_maps()
    products = _fetch_products(scope, scope_arg)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=PRODUCT_EXPORT_FIELDS)
    writer.writeheader()
    for p in products:
        row = {
            "id": p.get("id", ""),
            "name": p.get("name", ""),
            "description": p.get("description", ""),
            "category": cat_map.get(p.get("category_id"), ""),
            "subcategory": sub_map.get(p.get("subcategory_id"), ""),
            "price": p.get("price", ""),
            "sale_price": p.get("sale_price", "") or "",
            "currency": p.get("currency", "USD"),
            "product_type": p.get("product_type_name", ""),
            "stock_count": p.get("stock_count", 0),
            "is_active": p.get("is_active", True),
            "is_featured": p.get("is_featured", False),
            "delivery_note": p.get("delivery_note", "") or "",
            "warranty_info": p.get("warranty_info", "") or "",
            "min_quantity": p.get("min_quantity", "") or "",
            "max_quantity": p.get("max_quantity", "") or "",
            "bulk_purchase_enabled": p.get("bulk_purchase_enabled", True),
            "reusable": p.get("reusable", False),
            "product_emoji": p.get("product_emoji", "") or "",
            "sort_order": p.get("sort_order", "") or "",
            "created_at": p.get("created_at_str", ""),
        }
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def export_products_json(scope: str = "all", scope_arg: Optional[str] = None) -> bytes:
    cat_map, sub_map = _build_cat_maps()
    products = _fetch_products(scope, scope_arg)
    out = []
    for p in products:
        out.append({
            "id": p.get("id"),
            "name": p.get("name", ""),
            "description": p.get("description", ""),
            "category": cat_map.get(p.get("category_id"), ""),
            "subcategory": sub_map.get(p.get("subcategory_id"), ""),
            "price": p.get("price"),
            "sale_price": p.get("sale_price"),
            "currency": p.get("currency", "USD"),
            "product_type": p.get("product_type_name", ""),
            "stock_count": p.get("stock_count", 0),
            "is_active": p.get("is_active", True),
            "is_featured": p.get("is_featured", False),
            "delivery_note": p.get("delivery_note"),
            "warranty_info": p.get("warranty_info"),
            "min_quantity": p.get("min_quantity"),
            "max_quantity": p.get("max_quantity"),
            "bulk_purchase_enabled": p.get("bulk_purchase_enabled", True),
            "reusable": p.get("reusable", False),
            "product_emoji": p.get("product_emoji"),
            "sort_order": p.get("sort_order"),
            "created_at": p.get("created_at_str", ""),
        })
    return json.dumps({"products": out, "exported_at": datetime.utcnow().isoformat()},
                      ensure_ascii=False, indent=2).encode("utf-8")


def export_products_xlsx(scope: str = "all", scope_arg: Optional[str] = None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed — cannot export to Excel")

    cat_map, sub_map = _build_cat_maps()
    products = _fetch_products(scope, scope_arg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2196F3")

    ws.append(PRODUCT_EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for p in products:
        ws.append([
            p.get("id", ""),
            p.get("name", ""),
            p.get("description", ""),
            cat_map.get(p.get("category_id"), ""),
            sub_map.get(p.get("subcategory_id"), ""),
            p.get("price", ""),
            p.get("sale_price", "") or "",
            p.get("currency", "USD"),
            p.get("product_type_name", ""),
            p.get("stock_count", 0),
            p.get("is_active", True),
            p.get("is_featured", False),
            p.get("delivery_note", "") or "",
            p.get("warranty_info", "") or "",
            p.get("min_quantity", "") or "",
            p.get("max_quantity", "") or "",
            p.get("bulk_purchase_enabled", True),
            p.get("reusable", False),
            p.get("product_emoji", "") or "",
            p.get("sort_order", "") or "",
            p.get("created_at_str", ""),
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_products(
    file_format: str,
    scope: str = "all",
    scope_arg: Optional[str] = None,
    admin_id: int = 0,
) -> Tuple[bytes, int]:
    """Export products and record the export. Returns (bytes, row_count)."""
    from database import get_db_session
    from database.models import BulkExportRecord

    fmt = file_format.lower().strip(".")
    if fmt == "csv":
        data = export_products_csv(scope, scope_arg)
    elif fmt in ("xlsx", "xls"):
        data = export_products_xlsx(scope, scope_arg)
    elif fmt == "json":
        data = export_products_json(scope, scope_arg)
    else:
        raise ValueError(f"Unsupported format: {file_format}")

    # Count rows (approximate from CSV/JSON/XLSX)
    row_count = len(_fetch_products(scope, scope_arg))

    try:
        with get_db_session() as s:
            rec = BulkExportRecord(
                admin_id=admin_id,
                export_type="products",
                file_format=fmt,
                scope=scope,
                status="COMPLETED",
                row_count=row_count,
                size_bytes=len(data),
                started_at=datetime.utcnow(),
                completed_at=datetime.utcnow(),
            )
            s.add(rec)
    except Exception:
        logger.exception("Could not save BulkExportRecord")

    return data, row_count


# ─── Bulk actions on products ──────────────────────────────────────────────

def _product_ids_for_scope(
    scope: str,
    scope_arg: Optional[str],
    session,
) -> List[int]:
    from database.models import Product
    q = session.query(Product.id)
    if scope == "all":
        pass
    elif scope == "category" and scope_arg:
        try:
            q = q.filter(Product.category_id == int(scope_arg))
        except ValueError:
            return []
    elif scope == "selected" and scope_arg:
        try:
            return [int(x) for x in scope_arg.split(",") if x.strip()]
        except ValueError:
            return []
    elif scope == "active":
        q = q.filter(Product.is_active == True)
    elif scope == "inactive":
        q = q.filter(Product.is_active == False)
    return [r[0] for r in q.all()]


def _log_bulk_action(admin_id: int, action_type: str, entity_type: str, scope: str,
                     target_count: int, success_count: int, failed_count: int,
                     details: Optional[Dict] = None) -> None:
    from database import get_db_session
    from database.models import BulkActionRecord
    try:
        with get_db_session() as s:
            s.add(BulkActionRecord(
                admin_id=admin_id,
                action_type=action_type,
                entity_type=entity_type,
                scope=scope,
                target_count=target_count,
                success_count=success_count,
                failed_count=failed_count,
                details=json.dumps(details) if details else None,
                status="COMPLETED",
                created_at=datetime.utcnow(),
                completed_at=datetime.utcnow(),
            ))
    except Exception:
        logger.exception("Could not log BulkActionRecord for action=%s", action_type)


def bulk_enable_products(admin_id: int, scope: str, scope_arg: Optional[str] = None) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.is_active = True
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_enable", "product", scope, len(ids),
                     result["success"], result["failed"])
    return result


def bulk_disable_products(admin_id: int, scope: str, scope_arg: Optional[str] = None) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.is_active = False
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_disable", "product", scope, len(ids),
                     result["success"], result["failed"])
    return result


def bulk_edit_price(
    admin_id: int,
    scope: str,
    new_price: float,
    scope_arg: Optional[str] = None,
) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.price = new_price
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_price_edit", "product", scope, len(ids),
                     result["success"], result["failed"], {"new_price": new_price})
    return result


def bulk_edit_stock(
    admin_id: int,
    scope: str,
    new_stock: int,
    scope_arg: Optional[str] = None,
) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.stock_count = new_stock
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_stock_edit", "product", scope, len(ids),
                     result["success"], result["failed"], {"new_stock": new_stock})
    return result


def bulk_change_category(
    admin_id: int,
    scope: str,
    category_id: int,
    scope_arg: Optional[str] = None,
) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.category_id = category_id
                    p.subcategory_id = None
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_change_category", "product", scope, len(ids),
                     result["success"], result["failed"], {"category_id": category_id})
    return result


def bulk_update_tags(
    admin_id: int,
    scope: str,
    delivery_type: Optional[str] = None,
    scope_arg: Optional[str] = None,
) -> Dict:
    """Bulk update delivery type (product_type) for selected products."""
    from database import get_db_session
    from database.models import Product, ProductType
    result = {"success": 0, "failed": 0}
    if not delivery_type or delivery_type.upper() not in VALID_PRODUCT_TYPES:
        return {"success": 0, "failed": 0, "error": "Invalid delivery type"}
    try:
        new_ptype = ProductType[delivery_type.upper()]
    except KeyError:
        return {"success": 0, "failed": 0, "error": "Invalid delivery type"}

    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    p.product_type = new_ptype
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_update_delivery_type", "product", scope, len(ids),
                     result["success"], result["failed"], {"delivery_type": delivery_type})
    return result


def bulk_delete_products(
    admin_id: int,
    scope: str,
    scope_arg: Optional[str] = None,
) -> Dict:
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    s.delete(p)
                    result["success"] += 1
            except Exception:
                logger.exception("bulk_delete product %s failed", pid)
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_delete", "product", scope, len(ids),
                     result["success"], result["failed"])
    return result


def bulk_clone_products(
    admin_id: int,
    scope: str,
    scope_arg: Optional[str] = None,
) -> Dict:
    """Clone selected products (creates copies without keys)."""
    from database import get_db_session
    from database.models import Product
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        ids = _product_ids_for_scope(scope, scope_arg, s)
        # Batch-fetch all target products in a single query instead of one
        # SELECT per id (N+1) — significant at 1000+ product scale.
        _products = (
            {p.id: p for p in s.query(Product).filter(Product.id.in_(ids)).all()}
            if ids else {}
        )
        for pid in ids:
            try:
                p = _products.get(pid)
                if p:
                    clone = Product(
                        name=f"[Copy] {p.name}",
                        description=p.description,
                        price=p.price,
                        sale_price=p.sale_price,
                        stock_count=0,
                        product_type=p.product_type,
                        category_id=p.category_id,
                        subcategory_id=p.subcategory_id,
                        currency=p.currency,
                        is_active=False,
                        is_featured=False,
                        delivery_note=p.delivery_note,
                        warranty_info=p.warranty_info,
                        min_quantity=p.min_quantity,
                        max_quantity=p.max_quantity,
                        bulk_purchase_enabled=p.bulk_purchase_enabled,
                        reusable=p.reusable,
                        product_emoji=p.product_emoji,
                        delivery_format_template=p.delivery_format_template,
                        created_at=datetime.utcnow(),
                    )
                    s.add(clone)
                    result["success"] += 1
            except Exception:
                logger.exception("bulk_clone product %s failed", pid)
                result["failed"] += 1
    _log_bulk_action(admin_id, "bulk_clone", "product", scope, len(ids),
                     result["success"], result["failed"])
    return result


# ─── Import/Export statistics ──────────────────────────────────────────────

def get_product_bulk_stats() -> Dict[str, int]:
    from database import get_db_session
    from database.models import BulkImportRecord, BulkExportRecord, BulkActionRecord
    try:
        with get_db_session() as s:
            total_imports = s.query(BulkImportRecord).count()
            total_exports = s.query(BulkExportRecord).filter_by(export_type="products").count()
            total_imported = s.query(BulkImportRecord).with_entities(
                __import__("sqlalchemy").func.sum(BulkImportRecord.imported)
            ).scalar() or 0
            total_failed = s.query(BulkImportRecord).with_entities(
                __import__("sqlalchemy").func.sum(BulkImportRecord.failed)
            ).scalar() or 0
            bulk_actions = s.query(BulkActionRecord).filter_by(entity_type="product").count()
            return {
                "total_imports": total_imports,
                "total_exports": total_exports,
                "imported_products": int(total_imported),
                "failed_imports": int(total_failed),
                "bulk_actions": bulk_actions,
            }
    except Exception:
        logger.exception("get_product_bulk_stats failed")
        return {
            "total_imports": 0, "total_exports": 0,
            "imported_products": 0, "failed_imports": 0, "bulk_actions": 0,
        }
