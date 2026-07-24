"""Bulk User Management service — V35.

Handles user filtering, bulk actions, and CSV/Excel/JSON export.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import and_, or_

logger = logging.getLogger(__name__)

# ─── Export field list ─────────────────────────────────────────────────────

USER_EXPORT_FIELDS = [
    "id", "telegram_id", "username", "wallet_balance",
    "is_banned", "has_purchased", "loyalty_points",
    "preferred_currency", "referral_earnings",
    "last_seen_at", "created_at",
    "total_orders", "total_spent",
]

# ─── User filter helpers ───────────────────────────────────────────────────

def _apply_filter(q, filter_type: str, session):
    """Apply a named user filter to a SQLAlchemy query."""
    from database.models import User, Order, OrderStatus
    now = datetime.utcnow()
    inactive_cutoff = now - timedelta(days=30)
    vip_threshold = 500.0  # wallet balance or spent > this → VIP

    if filter_type == "active":
        q = q.filter(User.last_seen_at >= inactive_cutoff)
    elif filter_type == "inactive":
        q = q.filter(
            or_(User.last_seen_at < inactive_cutoff, User.last_seen_at == None)
        )
    elif filter_type == "banned":
        q = q.filter(User.is_banned == True)
    elif filter_type == "verified":
        q = q.filter(User.has_purchased == True)
    elif filter_type == "with_orders":
        q = q.filter(User.has_purchased == True)
    elif filter_type == "without_orders":
        q = q.filter(User.has_purchased == False)
    elif filter_type == "vip":
        q = q.filter(User.wallet_balance >= vip_threshold)
    elif filter_type == "all":
        pass
    return q


def search_users(
    search_by: str,
    search_value: str,
    filter_type: str = "all",
    limit: int = 50,
    offset: int = 0,
) -> Tuple[List[Dict], int]:
    """
    Search users by field + optional filter. Returns (rows, total_count).
    """
    from database import get_db_session
    from database.models import User

    results: List[Dict] = []
    total = 0
    try:
        with get_db_session() as s:
            q = s.query(User)

            # Apply search field
            if search_by == "telegram_id" and search_value:
                try:
                    q = q.filter(User.telegram_id == int(search_value))
                except ValueError:
                    return [], 0
            elif search_by == "username" and search_value:
                q = q.filter(User.username.ilike(f"%{search_value}%"))
            elif search_by == "name" and search_value:
                q = q.filter(User.username.ilike(f"%{search_value}%"))
            elif search_by == "min_balance" and search_value:
                try:
                    q = q.filter(User.wallet_balance >= float(search_value))
                except ValueError:
                    pass
            # filter_type applied on top
            q = _apply_filter(q, filter_type, s)
            total = q.count()
            users = q.order_by(User.created_at.desc()).limit(limit).offset(offset).all()

            for u in users:
                results.append({
                    "id": u.id,
                    "telegram_id": u.telegram_id,
                    "username": u.username or "—",
                    "wallet_balance": u.wallet_balance,
                    "is_banned": u.is_banned,
                    "has_purchased": u.has_purchased,
                    "loyalty_points": u.loyalty_points,
                    "preferred_currency": u.preferred_currency,
                    "referral_earnings": u.referral_earnings,
                    "last_seen_at": u.last_seen_at.strftime("%Y-%m-%d %H:%M") if u.last_seen_at else "—",
                    "created_at": u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "—",
                })
    except Exception:
        logger.exception("search_users failed")
    return results, total


def get_filtered_user_ids(filter_type: str, search_value: str = "") -> List[int]:
    """Return all user IDs matching a filter (for bulk actions)."""
    from database import get_db_session
    from database.models import User
    try:
        with get_db_session() as s:
            q = s.query(User.id)
            q = _apply_filter(q, filter_type, s)
            return [r[0] for r in q.all()]
    except Exception:
        logger.exception("get_filtered_user_ids failed")
        return []


def _log_bulk_user_action(
    admin_id: int,
    action_type: str,
    scope: str,
    target_count: int,
    success_count: int,
    failed_count: int,
    details: Optional[Dict] = None,
) -> None:
    from database import get_db_session
    from database.models import BulkActionRecord
    try:
        with get_db_session() as s:
            s.add(BulkActionRecord(
                admin_id=admin_id,
                action_type=action_type,
                entity_type="user",
                scope=scope,
                target_count=target_count,
                success_count=success_count,
                failed_count=failed_count,
                details=json.dumps(details) if details else None,
                status="COMPLETED",
                created_at=datetime.utcnow(),
                completed_at=datetime.utcnow(),
            ))
    except Exception:
        logger.exception("Could not log bulk user action %s", action_type)


# ─── Bulk actions ──────────────────────────────────────────────────────────

def bulk_ban_users(admin_id: int, user_ids: List[int]) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.is_banned = True
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_ban", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_unban_users(admin_id: int, user_ids: List[int]) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.is_banned = False
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_unban", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_add_balance(admin_id: int, user_ids: List[int], amount: float) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u and amount > 0:
                    u.wallet_balance = (u.wallet_balance or 0.0) + amount
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_add_balance", "selected", len(user_ids),
                          result["success"], result["failed"], {"amount": amount})
    return result


def bulk_remove_balance(admin_id: int, user_ids: List[int], amount: float) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u and amount > 0:
                    u.wallet_balance = max(0.0, (u.wallet_balance or 0.0) - amount)
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_remove_balance", "selected", len(user_ids),
                          result["success"], result["failed"], {"amount": amount})
    return result


def bulk_reset_wallet(admin_id: int, user_ids: List[int]) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.wallet_balance = 0.0
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_reset_wallet", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_reset_referral(admin_id: int, user_ids: List[int]) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.referral_earnings = 0.0
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_reset_referral", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_verify_users(admin_id: int, user_ids: List[int]) -> Dict:
    """Mark users as verified (has_purchased = True)."""
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.has_purchased = True
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_verify", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_unverify_users(admin_id: int, user_ids: List[int]) -> Dict:
    from database import get_db_session
    from database.models import User
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                u = s.query(User).filter_by(id=uid).first()
                if u:
                    u.has_purchased = False
                    result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_unverify", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


def bulk_delete_inactive_users(admin_id: int) -> Dict:
    """Delete users inactive for 90+ days with no orders and zero balance."""
    from database import get_db_session
    from database.models import User
    cutoff = datetime.utcnow() - timedelta(days=90)
    result = {"success": 0, "failed": 0, "skipped": 0}
    with get_db_session() as s:
        candidates = (
            s.query(User)
            .filter(
                User.has_purchased == False,
                User.wallet_balance <= 0,
                User.is_banned == False,
                or_(User.last_seen_at < cutoff, User.last_seen_at == None),
            )
            .all()
        )
        for u in candidates:
            try:
                s.delete(u)
                result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_delete_inactive", "inactive_90d",
                          result["success"] + result["failed"],
                          result["success"], result["failed"])
    return result


def bulk_reset_coupons(admin_id: int, user_ids: List[int]) -> Dict:
    """Remove all coupon redemptions for selected users."""
    from database import get_db_session
    from database.models import CouponRedemption
    result = {"success": 0, "failed": 0}
    with get_db_session() as s:
        for uid in user_ids:
            try:
                deleted = (
                    s.query(CouponRedemption)
                    .filter_by(user_id=uid)
                    .delete(synchronize_session="fetch")
                )
                result["success"] += 1
            except Exception:
                result["failed"] += 1
    _log_bulk_user_action(admin_id, "bulk_reset_coupons", "selected", len(user_ids),
                          result["success"], result["failed"])
    return result


# ─── User export ───────────────────────────────────────────────────────────

def _fetch_users_for_export(filter_type: str = "all") -> List[Dict]:
    from database import get_db_session
    from database.models import User, Order
    from sqlalchemy import func
    rows: List[Dict] = []
    try:
        with get_db_session() as s:
            # Subquery: order count and total spent per user
            order_stats = (
                s.query(
                    Order.user_id,
                    func.count(Order.id).label("total_orders"),
                    func.sum(Order.total_price).label("total_spent"),
                )
                .group_by(Order.user_id)
                .subquery()
            )

            q = s.query(User, order_stats.c.total_orders, order_stats.c.total_spent).outerjoin(
                order_stats, User.id == order_stats.c.user_id
            )
            q = _apply_filter(q, filter_type, s)

            for user, total_orders, total_spent in q.all():
                rows.append({
                    "id": user.id,
                    "telegram_id": user.telegram_id,
                    "username": user.username or "",
                    "wallet_balance": user.wallet_balance,
                    "is_banned": user.is_banned,
                    "has_purchased": user.has_purchased,
                    "loyalty_points": user.loyalty_points,
                    "preferred_currency": user.preferred_currency,
                    "referral_earnings": user.referral_earnings,
                    "last_seen_at": user.last_seen_at.strftime("%Y-%m-%d %H:%M:%S")
                                    if user.last_seen_at else "",
                    "created_at": user.created_at.strftime("%Y-%m-%d %H:%M:%S")
                                  if user.created_at else "",
                    "total_orders": int(total_orders or 0),
                    "total_spent": float(total_spent or 0),
                })
    except Exception:
        logger.exception("_fetch_users_for_export failed")
    return rows


def export_users_csv(filter_type: str = "all") -> bytes:
    rows = _fetch_users_for_export(filter_type)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=USER_EXPORT_FIELDS)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def export_users_json(filter_type: str = "all") -> bytes:
    rows = _fetch_users_for_export(filter_type)
    return json.dumps(
        {"users": rows, "exported_at": datetime.utcnow().isoformat(), "filter": filter_type},
        ensure_ascii=False, indent=2,
    ).encode("utf-8")


def export_users_xlsx(filter_type: str = "all") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    rows = _fetch_users_for_export(filter_type)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4CAF50")

    ws.append(USER_EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(f, "") for f in USER_EXPORT_FIELDS])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_users(
    file_format: str,
    filter_type: str = "all",
    admin_id: int = 0,
) -> Tuple[bytes, int]:
    """Export users and record the export. Returns (bytes, row_count)."""
    from database import get_db_session
    from database.models import BulkExportRecord

    fmt = file_format.lower().strip(".")
    if fmt == "csv":
        data = export_users_csv(filter_type)
    elif fmt in ("xlsx", "xls"):
        data = export_users_xlsx(filter_type)
    elif fmt == "json":
        data = export_users_json(filter_type)
    else:
        raise ValueError(f"Unsupported format: {file_format}")

    row_count = len(_fetch_users_for_export(filter_type))

    try:
        with get_db_session() as s:
            s.add(BulkExportRecord(
                admin_id=admin_id,
                export_type="users",
                file_format=fmt,
                scope=filter_type,
                status="COMPLETED",
                row_count=row_count,
                size_bytes=len(data),
                started_at=datetime.utcnow(),
                completed_at=datetime.utcnow(),
            ))
    except Exception:
        logger.exception("Could not save BulkExportRecord (users)")

    return data, row_count


# ─── Statistics ────────────────────────────────────────────────────────────

def get_user_bulk_stats() -> Dict[str, int]:
    from database import get_db_session
    from database.models import BulkExportRecord, BulkActionRecord, User
    try:
        with get_db_session() as s:
            total_exports = s.query(BulkExportRecord).filter_by(export_type="users").count()
            bulk_actions = s.query(BulkActionRecord).filter_by(entity_type="user").count()
            managed = s.query(BulkActionRecord).filter_by(entity_type="user").with_entities(
                __import__("sqlalchemy").func.sum(BulkActionRecord.success_count)
            ).scalar() or 0
            total_users = s.query(User).count()
            return {
                "total_exports": total_exports,
                "bulk_user_actions": bulk_actions,
                "total_managed_users": int(managed),
                "total_users": total_users,
            }
    except Exception:
        logger.exception("get_user_bulk_stats failed")
        return {
            "total_exports": 0, "bulk_user_actions": 0,
            "total_managed_users": 0, "total_users": 0,
        }
