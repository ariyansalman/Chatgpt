"""Enterprise Broadcast Analytics Service.

Provides:
  • get_analytics_dashboard()      — aggregate admin dashboard stats
  • get_broadcast_analytics()      — per-broadcast real-time stats
  • search_broadcast_history()     — search + filter + paginate history
  • get_error_breakdown()          — categorise errors in retry queue
  • generate_*_report()            — typed report dicts (delivery, failure, …)
  • generate_period_report()       — daily / weekly / monthly summary
  • export_csv / excel / json / pdf — bytes for file attachments
  • retry_failed_deliveries()      — re-queue failed items for the scheduler
  • clear_retry_queue()            — clear pending retries for a broadcast
  • archive_broadcast()            — mark archived
  • delete_broadcast_history()     — hard-delete a broadcast record
  • log_export()                   — write to broadcast_export_history
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from database import get_db_session
from database.models import ScheduledBroadcast, BroadcastLog, BroadcastRetryQueue
from utils.bot_config import cfg

logger = logging.getLogger(__name__)

# ── Error classification ──────────────────────────────────────────────────────

_ERROR_PATTERNS: List[Tuple[str, str]] = [
    (r"bot was blocked",                             "Blocked Users"),
    (r"user is deactivated",                         "Invalid Users"),
    (r"chat not found|peer_id_invalid",              "Invalid Users"),
    (r"flood|retry after|too many requests",         "Flood Wait"),
    (r"bad request|forbidden|unauthorized",          "API Errors"),
    (r"timeout|network|connection",                  "Network Errors"),
    (r"message is too long|media.*invalid",          "Telegram Errors"),
]


def _classify_error(msg: Optional[str]) -> str:
    if not msg:
        return "Unknown Errors"
    low = msg.lower()
    for pattern, label in _ERROR_PATTERNS:
        if re.search(pattern, low):
            return label
    return "Unknown Errors"


# ── Internal helpers ──────────────────────────────────────────────────────────

def _order_stats_agg(s) -> Tuple[int, int, int]:
    """Return (total_sent, total_delivered, total_failed) across all logs."""
    from sqlalchemy import func
    row = s.query(
        func.coalesce(func.sum(BroadcastLog.sent),      0),
        func.coalesce(func.sum(BroadcastLog.delivered), 0),
        func.coalesce(func.sum(BroadcastLog.failed),    0),
    ).first()
    return int(row[0]), int(row[1]), int(row[2])


def _avg_speed(s) -> Optional[float]:
    """Return average messages/second across completed log runs."""
    from sqlalchemy import func
    logs = (s.query(BroadcastLog)
             .filter(BroadcastLog.started_at.isnot(None),
                     BroadcastLog.finished_at.isnot(None),
                     BroadcastLog.sent > 0)
             .order_by(BroadcastLog.created_at.desc())
             .limit(100)
             .all())
    if not logs:
        return None
    speeds = []
    for l in logs:
        secs = (l.finished_at - l.started_at).total_seconds()
        if secs > 0:
            speeds.append(l.sent / secs)
    return round(sum(speeds) / len(speeds), 2) if speeds else None


def _br_info(br: ScheduledBroadcast) -> Dict[str, Any]:
    """Serialize a ScheduledBroadcast to a plain dict for reporting."""
    return {
        "id":           br.id,
        "title":        br.title,
        "broadcast_type": getattr(br, "broadcast_type", None),
        "status":       br.status,
        "media_type":   br.media_type,
        "target_segment": br.target_segment,
        "created_by":   br.created_by,
        "created_at":   br.created_at.isoformat()   if br.created_at   else None,
        "started_at":   br.started_at.isoformat()   if br.started_at   else None,
        "finished_at":  br.finished_at.isoformat()  if br.finished_at  else None,
        "scheduled_at": br.scheduled_at.isoformat() if br.scheduled_at else None,
        "is_recurring": br.is_recurring,
        "recurrence_type": br.recurrence_type,
        "total_recipients": br.total_recipients or 0,
        "sent_count":      br.sent_count      or 0,
        "delivered_count": br.delivered_count or 0,
        "failed_count":    br.failed_count    or 0,
        "blocked_count":   br.blocked_count   or 0,
        "skipped_count":   br.skipped_count   or 0,
    }


# ── Dashboard ─────────────────────────────────────────────────────────────────

def get_analytics_dashboard() -> Dict[str, Any]:
    """Aggregate stats for the Admin Analytics Dashboard."""
    now         = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    with get_db_session() as s:
        from sqlalchemy import func

        total     = s.query(ScheduledBroadcast).count()
        today_bc  = s.query(ScheduledBroadcast).filter(
            ScheduledBroadcast.created_at >= today_start).count()
        running   = s.query(ScheduledBroadcast).filter_by(status="sending").count()
        completed = s.query(ScheduledBroadcast).filter_by(status="sent").count()
        failed_bc = s.query(ScheduledBroadcast).filter_by(status="failed").count()
        paused    = s.query(ScheduledBroadcast).filter_by(status="paused").count()
        scheduled = s.query(ScheduledBroadcast).filter_by(status="scheduled").count()
        cancelled = s.query(ScheduledBroadcast).filter_by(status="cancelled").count()

        total_sent, total_delivered, total_failed = _order_stats_agg(s)
        success_rate = round(total_delivered / total_sent * 100, 1) if total_sent else 0.0
        failure_rate = round(total_failed    / total_sent * 100, 1) if total_sent else 0.0

        avg_speed = _avg_speed(s)

        retry_pending = s.query(BroadcastRetryQueue).filter_by(status="pending").count()
        blocked_total = (s.query(BroadcastRetryQueue)
                         .filter(BroadcastRetryQueue.error_msg.ilike("%blocked%"))
                         .count())

        # Average delivery time (seconds per message)
        avg_delivery_s: Optional[float] = None
        logs_timed = (s.query(BroadcastLog)
                       .filter(BroadcastLog.started_at.isnot(None),
                               BroadcastLog.finished_at.isnot(None),
                               BroadcastLog.sent > 0)
                       .order_by(BroadcastLog.created_at.desc())
                       .limit(50)
                       .all())
        if logs_timed:
            durs = [(l.finished_at - l.started_at).total_seconds() / l.sent
                    for l in logs_timed
                    if (l.finished_at - l.started_at).total_seconds() > 0]
            if durs:
                avg_delivery_s = round(sum(durs) / len(durs) * 1000, 1)  # ms

    return {
        "total":          total,
        "today":          today_bc,
        "running":        running,
        "completed":      completed,
        "failed":         failed_bc,
        "paused":         paused,
        "scheduled":      scheduled,
        "cancelled":      cancelled,
        "total_sent":     total_sent,
        "total_delivered": total_delivered,
        "total_failed":   total_failed,
        "success_rate":   success_rate,
        "failure_rate":   failure_rate,
        "avg_speed_mps":  avg_speed,       # messages/second
        "avg_delivery_ms": avg_delivery_s,
        "retry_pending":  retry_pending,
        "blocked_total":  blocked_total,
    }


# ── Per-broadcast analytics ───────────────────────────────────────────────────

def get_broadcast_analytics(broadcast_id: int) -> Dict[str, Any]:
    """Full real-time analytics for a single broadcast."""
    with get_db_session() as s:
        br = s.get(ScheduledBroadcast, broadcast_id)
        if not br:
            return {}

        logs = (s.query(BroadcastLog)
                 .filter_by(broadcast_id=broadcast_id)
                 .order_by(BroadcastLog.created_at.desc())
                 .all())

        retry_pending = (s.query(BroadcastRetryQueue)
                          .filter_by(broadcast_id=broadcast_id, status="pending").count())
        retry_sent    = (s.query(BroadcastRetryQueue)
                          .filter_by(broadcast_id=broadcast_id, status="sent").count())
        retry_failed  = (s.query(BroadcastRetryQueue)
                          .filter_by(broadcast_id=broadcast_id, status="failed").count())

        sent      = br.sent_count      or 0
        delivered = br.delivered_count or 0
        failed    = br.failed_count    or 0
        blocked   = br.blocked_count   or 0
        skipped   = br.skipped_count   or 0
        total     = br.total_recipients or 0
        remaining = max(0, total - sent)

        success_rate = round(delivered / sent * 100, 1) if sent else 0.0
        failure_rate = round(failed    / sent * 100, 1) if sent else 0.0

        # Delivery speed (msg/s) from latest log
        avg_speed = None
        if logs:
            last = logs[0]
            if last.started_at and last.finished_at and last.sent:
                secs = (last.finished_at - last.started_at).total_seconds()
                if secs > 0:
                    avg_speed = round(last.sent / secs, 2)

        # Elapsed / ETA
        elapsed_s = None
        eta_s     = None
        if br.started_at:
            elapsed_s = (datetime.utcnow() - br.started_at).total_seconds()
            if avg_speed and remaining > 0:
                eta_s = remaining / avg_speed

        run_logs = []
        for l in logs:
            dur = None
            if l.started_at and l.finished_at:
                dur = round((l.finished_at - l.started_at).total_seconds(), 1)
            run_logs.append({
                "run_at":    l.created_at.isoformat() if l.created_at else None,
                "total":     l.total_recipients,
                "sent":      l.sent,
                "delivered": l.delivered,
                "failed":    l.failed,
                "blocked":   l.blocked,
                "skipped":   l.skipped,
                "duration_s": dur,
                "speed_mps":  round(l.sent / dur, 2) if dur and dur > 0 and l.sent else None,
            })

        return {
            **_br_info(br),
            "remaining":      remaining,
            "success_rate":   success_rate,
            "failure_rate":   failure_rate,
            "avg_speed_mps":  avg_speed,
            "elapsed_s":      round(elapsed_s, 0) if elapsed_s else None,
            "eta_s":          round(eta_s,     0) if eta_s     else None,
            "retry_pending":  retry_pending,
            "retry_sent":     retry_sent,
            "retry_failed":   retry_failed,
            "run_count":      len(logs),
            "run_logs":       run_logs,
        }


# ── History search / filter ───────────────────────────────────────────────────

def search_broadcast_history(
    query: Optional[str] = None,
    filter_status: Optional[str] = None,
    filter_type: Optional[str] = None,
    page: int = 0,
    page_size: int = 10,
    include_archived: bool = False,
) -> Tuple[List[Dict[str, Any]], int]:
    """Return (records, total_count) with search / filter / pagination."""
    max_history = cfg.get_int("broadcast_max_history", 500)
    with get_db_session() as s:
        q = s.query(ScheduledBroadcast)
        if not include_archived:
            try:
                q = q.filter(ScheduledBroadcast.is_archived == False)
            except Exception:
                pass
        if filter_status and filter_status != "all":
            q = q.filter(ScheduledBroadcast.status == filter_status)
        if filter_type and filter_type != "all":
            q = q.filter(ScheduledBroadcast.broadcast_type == filter_type)
        if query:
            like = f"%{query}%"
            q = q.filter(
                ScheduledBroadcast.title.ilike(like) |
                ScheduledBroadcast.target_segment.ilike(like)
            )
        total = min(q.count(), max_history)
        rows  = (q.order_by(ScheduledBroadcast.created_at.desc())
                  .offset(page * page_size)
                  .limit(page_size)
                  .all())
        return [_br_info(r) for r in rows], total


# ── Error breakdown ───────────────────────────────────────────────────────────

def get_error_breakdown(broadcast_id: int) -> Dict[str, int]:
    """Return {error_category: count} for all retry-queue entries."""
    with get_db_session() as s:
        entries = (s.query(BroadcastRetryQueue)
                    .filter_by(broadcast_id=broadcast_id)
                    .all())
    tally: Dict[str, int] = {}
    for e in entries:
        cat = _classify_error(e.error_msg)
        tally[cat] = tally.get(cat, 0) + 1
    return tally


def get_error_detail(broadcast_id: int, category: Optional[str] = None) -> List[Dict[str, Any]]:
    """Return list of per-user error detail rows for the retry queue."""
    with get_db_session() as s:
        q = s.query(BroadcastRetryQueue).filter_by(broadcast_id=broadcast_id)
        rows = q.order_by(BroadcastRetryQueue.id.desc()).limit(200).all()
    result = []
    for r in rows:
        cat = _classify_error(r.error_msg)
        if category and cat != category:
            continue
        result.append({
            "id":          r.id,
            "telegram_id": r.telegram_id,
            "status":      r.status,
            "error_msg":   r.error_msg or "",
            "category":    cat,
            "attempts":    r.attempts,
            "retry_at":    r.retry_at.isoformat() if r.retry_at else None,
            "created_at":  r.created_at.isoformat() if getattr(r, "created_at", None) else None,
        })
    return result


# ── Report generators ─────────────────────────────────────────────────────────

def _base_report(broadcast_id: int) -> Dict[str, Any]:
    """Return base analytics dict for a broadcast (used by all typed reports)."""
    return get_broadcast_analytics(broadcast_id)


def generate_delivery_report(broadcast_id: int) -> Dict[str, Any]:
    base = _base_report(broadcast_id)
    base["report_type"] = "delivery"
    return base


def generate_failure_report(broadcast_id: int) -> Dict[str, Any]:
    base   = _base_report(broadcast_id)
    detail = get_error_detail(broadcast_id)
    failures = [r for r in detail if r["status"] in ("failed", "pending")]
    base["report_type"]  = "failure"
    base["failure_rows"] = failures
    base["error_breakdown"] = get_error_breakdown(broadcast_id)
    return base


def generate_blocked_report(broadcast_id: int) -> Dict[str, Any]:
    detail = get_error_detail(broadcast_id, category="Blocked Users")
    base   = _base_report(broadcast_id)
    base["report_type"]   = "blocked"
    base["blocked_rows"]  = detail
    return base


def generate_skipped_report(broadcast_id: int) -> Dict[str, Any]:
    base = _base_report(broadcast_id)
    base["report_type"] = "skipped"
    return base


def generate_success_report(broadcast_id: int) -> Dict[str, Any]:
    base = _base_report(broadcast_id)
    base["report_type"] = "success"
    detail = get_error_detail(broadcast_id)
    base["sent_rows"] = [r for r in detail if r["status"] == "sent"]
    return base


def generate_retry_report(broadcast_id: int) -> Dict[str, Any]:
    base   = _base_report(broadcast_id)
    detail = get_error_detail(broadcast_id)
    base["report_type"]  = "retry"
    base["retry_rows"]   = detail
    base["error_breakdown"] = get_error_breakdown(broadcast_id)
    return base


def generate_period_report(period: str) -> Dict[str, Any]:
    """Generate daily / weekly / monthly aggregate summary."""
    now = datetime.utcnow()
    if period == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        label = "Daily"
    elif period == "weekly":
        start = now - timedelta(days=7)
        label = "Weekly (last 7 days)"
    elif period == "monthly":
        start = now - timedelta(days=30)
        label = "Monthly (last 30 days)"
    else:
        start = now - timedelta(days=1)
        label = "Daily"

    with get_db_session() as s:
        from sqlalchemy import func
        brs = (s.query(ScheduledBroadcast)
                .filter(ScheduledBroadcast.created_at >= start)
                .all())
        rows = [_br_info(b) for b in brs]

        agg = s.query(
            func.coalesce(func.sum(BroadcastLog.sent),      0),
            func.coalesce(func.sum(BroadcastLog.delivered), 0),
            func.coalesce(func.sum(BroadcastLog.failed),    0),
            func.coalesce(func.sum(BroadcastLog.blocked),   0),
            func.coalesce(func.sum(BroadcastLog.skipped),   0),
        ).join(ScheduledBroadcast,
               BroadcastLog.broadcast_id == ScheduledBroadcast.id
               ).filter(ScheduledBroadcast.created_at >= start).first()

        status_counts: Dict[str, int] = {}
        for b in brs:
            status_counts[b.status] = status_counts.get(b.status, 0) + 1

    sent_total = int(agg[0])
    delivered  = int(agg[1])
    failed     = int(agg[2])
    blocked    = int(agg[3])
    skipped    = int(agg[4])

    return {
        "report_type":    "period",
        "period":         period,
        "label":          label,
        "from":           start.isoformat(),
        "to":             now.isoformat(),
        "broadcasts":     rows,
        "broadcast_count": len(rows),
        "status_counts":  status_counts,
        "total_sent":     sent_total,
        "total_delivered": delivered,
        "total_failed":   failed,
        "total_blocked":  blocked,
        "total_skipped":  skipped,
        "success_rate":   round(delivered / sent_total * 100, 1) if sent_total else 0.0,
        "failure_rate":   round(failed    / sent_total * 100, 1) if sent_total else 0.0,
    }


# ── Export ────────────────────────────────────────────────────────────────────

def export_csv(data: Dict[str, Any]) -> bytes:
    """Serialize report data to CSV bytes."""
    buf = io.StringIO()
    w   = csv.writer(buf)

    report_type = data.get("report_type", "report")
    w.writerow(["Report Type", report_type.capitalize()])
    w.writerow(["Generated At", datetime.utcnow().isoformat()])
    w.writerow([])

    # Header fields
    skip_keys = {"run_logs", "failure_rows", "blocked_rows", "sent_rows", "retry_rows",
                 "broadcasts", "error_breakdown", "status_counts", "report_type"}
    for k, v in data.items():
        if k not in skip_keys:
            w.writerow([k, v])
    w.writerow([])

    # Sub-tables
    if "error_breakdown" in data:
        w.writerow(["Error Breakdown"])
        w.writerow(["Category", "Count"])
        for cat, cnt in data["error_breakdown"].items():
            w.writerow([cat, cnt])
        w.writerow([])

    for table_key, headers in [
        ("run_logs",     ["Run At", "Total", "Sent", "Delivered", "Failed", "Blocked", "Skipped", "Duration (s)", "Speed (m/s)"]),
        ("failure_rows", ["Telegram ID", "Status", "Error", "Category", "Attempts", "Retry At"]),
        ("blocked_rows", ["Telegram ID", "Error", "Attempts", "Retry At"]),
        ("retry_rows",   ["Telegram ID", "Status", "Error", "Category", "Attempts", "Retry At"]),
        ("sent_rows",    ["Telegram ID", "Status", "Attempts"]),
        ("broadcasts",   ["ID", "Title", "Status", "Sent", "Delivered", "Failed", "Created At"]),
    ]:
        if table_key in data and data[table_key]:
            w.writerow([table_key.replace("_", " ").title()])
            w.writerow(headers)
            for row in data[table_key]:
                if table_key == "run_logs":
                    w.writerow([row.get("run_at"), row.get("total"), row.get("sent"),
                                 row.get("delivered"), row.get("failed"), row.get("blocked"),
                                 row.get("skipped"), row.get("duration_s"), row.get("speed_mps")])
                elif table_key in ("failure_rows", "retry_rows"):
                    w.writerow([row.get("telegram_id"), row.get("status"), row.get("error_msg"),
                                 row.get("category"), row.get("attempts"), row.get("retry_at")])
                elif table_key == "blocked_rows":
                    w.writerow([row.get("telegram_id"), row.get("error_msg"),
                                 row.get("attempts"), row.get("retry_at")])
                elif table_key == "sent_rows":
                    w.writerow([row.get("telegram_id"), row.get("status"), row.get("attempts")])
                elif table_key == "broadcasts":
                    w.writerow([row.get("id"), row.get("title"), row.get("status"),
                                 row.get("sent_count"), row.get("delivered_count"),
                                 row.get("failed_count"), row.get("created_at")])
            w.writerow([])

    if "status_counts" in data:
        w.writerow(["Status Breakdown"])
        w.writerow(["Status", "Count"])
        for st, cnt in data["status_counts"].items():
            w.writerow([st, cnt])

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def export_excel(data: Dict[str, Any], title: str = "Broadcast Report") -> bytes:
    """Serialize report data to Excel bytes (.xlsx) using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Summary"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    section_font = Font(bold=True, color="1F4E79")

    def _hrow(ws, values: List[str], row: int) -> None:
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font  = header_font
            c.fill  = header_fill
            c.alignment = Alignment(horizontal="center")

    def _srow(ws, label: str, row: int) -> None:
        c = ws.cell(row=row, column=1, value=label)
        c.font = section_font

    # Summary sheet
    row = 1
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    row += 2

    skip_keys = {"run_logs", "failure_rows", "blocked_rows", "sent_rows", "retry_rows",
                 "broadcasts", "error_breakdown", "status_counts", "report_type"}
    for k, v in data.items():
        if k not in skip_keys:
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v) if v is not None else "")
            row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30

    # Error breakdown sheet
    if "error_breakdown" in data and data["error_breakdown"]:
        ws2 = wb.create_sheet("Error Breakdown")
        _hrow(ws2, ["Error Category", "Count"], 1)
        for i, (cat, cnt) in enumerate(data["error_breakdown"].items(), 2):
            ws2.cell(row=i, column=1, value=cat)
            ws2.cell(row=i, column=2, value=cnt)
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 10

    # Run Logs sheet
    if "run_logs" in data and data["run_logs"]:
        ws3 = wb.create_sheet("Run Logs")
        headers = ["Run At", "Total", "Sent", "Delivered", "Failed", "Blocked", "Skipped", "Duration (s)", "Speed (m/s)"]
        _hrow(ws3, headers, 1)
        for i, row_d in enumerate(data["run_logs"], 2):
            ws3.cell(row=i, column=1, value=row_d.get("run_at"))
            ws3.cell(row=i, column=2, value=row_d.get("total"))
            ws3.cell(row=i, column=3, value=row_d.get("sent"))
            ws3.cell(row=i, column=4, value=row_d.get("delivered"))
            ws3.cell(row=i, column=5, value=row_d.get("failed"))
            ws3.cell(row=i, column=6, value=row_d.get("blocked"))
            ws3.cell(row=i, column=7, value=row_d.get("skipped"))
            ws3.cell(row=i, column=8, value=row_d.get("duration_s"))
            ws3.cell(row=i, column=9, value=row_d.get("speed_mps"))
        for col in range(1, 10):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # Delivery Failures sheet
    for sheet_key, sheet_name, cols in [
        ("failure_rows", "Failures",
         ["Telegram ID", "Status", "Error Message", "Category", "Attempts", "Retry At"]),
        ("retry_rows",   "Retry Queue",
         ["Telegram ID", "Status", "Error Message", "Category", "Attempts", "Retry At"]),
        ("broadcasts",   "Broadcasts",
         ["ID", "Title", "Status", "Sent", "Delivered", "Failed", "Created At"]),
    ]:
        if sheet_key in data and data[sheet_key]:
            ws_s = wb.create_sheet(sheet_name)
            _hrow(ws_s, cols, 1)
            for i, row_d in enumerate(data[sheet_key], 2):
                if sheet_key in ("failure_rows", "retry_rows"):
                    ws_s.cell(row=i, column=1, value=row_d.get("telegram_id"))
                    ws_s.cell(row=i, column=2, value=row_d.get("status"))
                    ws_s.cell(row=i, column=3, value=row_d.get("error_msg", "")[:500])
                    ws_s.cell(row=i, column=4, value=row_d.get("category"))
                    ws_s.cell(row=i, column=5, value=row_d.get("attempts"))
                    ws_s.cell(row=i, column=6, value=row_d.get("retry_at"))
                elif sheet_key == "broadcasts":
                    ws_s.cell(row=i, column=1, value=row_d.get("id"))
                    ws_s.cell(row=i, column=2, value=row_d.get("title"))
                    ws_s.cell(row=i, column=3, value=row_d.get("status"))
                    ws_s.cell(row=i, column=4, value=row_d.get("sent_count"))
                    ws_s.cell(row=i, column=5, value=row_d.get("delivered_count"))
                    ws_s.cell(row=i, column=6, value=row_d.get("failed_count"))
                    ws_s.cell(row=i, column=7, value=row_d.get("created_at"))
            for col in range(1, len(cols) + 1):
                ws_s.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_json(data: Dict[str, Any]) -> bytes:
    """Serialize report data to JSON bytes."""
    return json.dumps(data, indent=2, ensure_ascii=False, default=str).encode("utf-8")


def export_pdf(data: Dict[str, Any], title: str = "Broadcast Report") -> bytes:
    """Serialize report data to PDF bytes using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm,  bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, spaceAfter=4)
    body = styles["Normal"]

    story.append(Paragraph(title, h1))
    story.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", body))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 0.4*cm))

    skip_keys = {"run_logs", "failure_rows", "blocked_rows", "sent_rows", "retry_rows",
                 "broadcasts", "error_breakdown", "status_counts", "report_type"}

    # Summary table
    summary_rows = [["Field", "Value"]]
    for k, v in data.items():
        if k not in skip_keys:
            summary_rows.append([str(k).replace("_", " ").title(), str(v) if v is not None else "—"])
    if len(summary_rows) > 1:
        story.append(Paragraph("Summary", h2))
        t = Table(summary_rows, colWidths=[7*cm, 10*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))

    # Error breakdown
    if "error_breakdown" in data and data["error_breakdown"]:
        story.append(Paragraph("Error Breakdown", h2))
        err_rows = [["Category", "Count"]] + [
            [cat, str(cnt)] for cat, cnt in data["error_breakdown"].items()]
        t2 = Table(err_rows, colWidths=[12*cm, 5*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C0392B")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDEDEC")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
            ("LEFTPADDING",(0, 0), (-1, -1), 6),
        ]))
        story.append(t2)
        story.append(Spacer(1, 0.4*cm))

    # Run logs (up to 20 rows in PDF to keep it readable)
    if "run_logs" in data and data["run_logs"]:
        story.append(Paragraph("Run Logs", h2))
        hdr = ["Run At", "Sent", "Delivered", "Failed", "Duration (s)", "Speed"]
        rows = [hdr] + [
            [r.get("run_at", "")[:19], str(r.get("sent","")), str(r.get("delivered","")),
             str(r.get("failed","")), str(r.get("duration_s","")), str(r.get("speed_mps",""))]
            for r in data["run_logs"][:20]
        ]
        t3 = Table(rows, colWidths=[4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2ECC71")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAFAF1")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
            ("LEFTPADDING",(0, 0), (-1, -1), 4),
        ]))
        story.append(t3)

    doc.build(story)
    return buf.getvalue()


# ── Retry management ──────────────────────────────────────────────────────────

def retry_failed_deliveries(broadcast_id: int) -> int:
    """Re-queue all failed/stuck retry-queue entries. Returns count re-queued."""
    with get_db_session() as s:
        entries = (s.query(BroadcastRetryQueue)
                    .filter_by(broadcast_id=broadcast_id, status="failed")
                    .all())
        count = 0
        for e in entries:
            e.status   = "pending"
            e.retry_at = datetime.utcnow()
            count += 1
        s.commit()
    return count


def clear_retry_queue(broadcast_id: int) -> int:
    """Delete all pending retry entries. Returns count deleted."""
    with get_db_session() as s:
        entries = (s.query(BroadcastRetryQueue)
                    .filter_by(broadcast_id=broadcast_id, status="pending")
                    .all())
        count = len(entries)
        for e in entries:
            s.delete(e)
        s.commit()
    return count


# ── History management ────────────────────────────────────────────────────────

def archive_broadcast(broadcast_id: int) -> bool:
    """Mark a broadcast as archived. Returns True on success."""
    with get_db_session() as s:
        br = s.get(ScheduledBroadcast, broadcast_id)
        if not br:
            return False
        try:
            br.is_archived = True
        except Exception:
            pass  # column may not exist yet if migration not run
        br.updated_at = datetime.utcnow()
        s.commit()
    return True


def delete_broadcast_history(broadcast_id: int) -> bool:
    """Hard-delete a broadcast and all its logs. Returns True on success."""
    with get_db_session() as s:
        br = s.get(ScheduledBroadcast, broadcast_id)
        if not br:
            return False
        s.delete(br)
        s.commit()
    return True


# ── Export history logging ────────────────────────────────────────────────────

def log_export(
    broadcast_id: Optional[int],
    export_type: str,
    report_type: str,
    period: Optional[str],
    generated_by: Optional[int],
    file_size_bytes: int,
    filename: str,
) -> None:
    """Write an entry to broadcast_export_history (non-blocking, swallows errors)."""
    try:
        from sqlalchemy import text
        with get_db_session() as s:
            s.execute(
                text(
                    "INSERT INTO broadcast_export_history "
                    "(broadcast_id, export_type, report_type, period, "
                    " generated_at, generated_by, file_size_bytes, filename) "
                    "VALUES (:bid, :et, :rt, :p, :ga, :gb, :fsz, :fn)"
                ),
                {
                    "bid": broadcast_id,
                    "et":  export_type,
                    "rt":  report_type,
                    "p":   period,
                    "ga":  datetime.utcnow(),
                    "gb":  generated_by,
                    "fsz": file_size_bytes,
                    "fn":  filename,
                },
            )
            s.commit()
    except Exception:
        logger.debug("log_export: failed (non-critical)", exc_info=True)
