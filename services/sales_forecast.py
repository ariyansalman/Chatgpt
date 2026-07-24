"""Sales Forecast & Business Insights Service — V40.

Provides:
  • Revenue / order / customer metric aggregations
  • Simple moving-average (SMA) based sales forecasting
  • Business insights (best/worst sellers, top customers, etc.)
  • Product insights (low stock, slow sellers, trending)
  • Report generation (daily / weekly / monthly / yearly / themed)
  • Export to CSV, JSON, Excel (.xlsx), PDF
"""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timedelta, date
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import func as sqlfunc, desc, asc

from database import get_db_session
from database.models import (
    Order, OrderItem, Product, Category, User, Transaction,
    ReferralReward, PaymentMethod as PaymentMethodEnum,
    OrderStatus, BusinessReport, ForecastSnapshot, DailyAnalyticsSnapshot,
)

logger = logging.getLogger(__name__)

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _utc_today() -> datetime:
    now = datetime.utcnow()
    return now.replace(hour=0, minute=0, second=0, microsecond=0)


def _window(days_ago: int, days_span: int = 1) -> Tuple[datetime, datetime]:
    today = _utc_today()
    end   = today - timedelta(days=days_ago)
    start = end - timedelta(days=days_span)
    return start, end


# ─── Core Revenue Metrics ─────────────────────────────────────────────────────

def _revenue_in_range(start: datetime, end: datetime) -> float:
    with get_db_session() as s:
        result = (s.query(sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0.0))
                  .filter(Order.status.in_([
                      OrderStatus.COMPLETED.value,
                      OrderStatus.DELIVERED.value if hasattr(OrderStatus, 'DELIVERED') else OrderStatus.COMPLETED.value,
                  ]))
                  .filter(Order.created_at >= start, Order.created_at < end)
                  .scalar())
        return float(result or 0)


def _orders_in_range(start: datetime, end: datetime) -> int:
    with get_db_session() as s:
        return (s.query(Order)
                .filter(Order.status.in_([
                    OrderStatus.COMPLETED.value,
                ]))
                .filter(Order.created_at >= start, Order.created_at < end)
                .count())


def get_revenue_summary() -> Dict[str, float]:
    """Return key revenue figures: today, yesterday, weekly, monthly, yearly, total."""
    today = _utc_today()
    yesterday_start = today - timedelta(days=1)

    with get_db_session() as s:
        def _rev(start: datetime, end: datetime) -> float:
            r = (s.query(sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0.0))
                 .filter(Order.status == OrderStatus.COMPLETED.value)
                 .filter(Order.created_at >= start, Order.created_at < end)
                 .scalar())
            return float(r or 0)

        def _orders(start: datetime, end: datetime) -> int:
            return (s.query(sqlfunc.count(Order.id))
                    .filter(Order.status == OrderStatus.COMPLETED.value)
                    .filter(Order.created_at >= start, Order.created_at < end)
                    .scalar() or 0)

        now = datetime.utcnow()
        week_start  = today - timedelta(days=7)
        month_start = today - timedelta(days=30)
        year_start  = today - timedelta(days=365)

        today_rev    = _rev(today, now)
        yest_rev     = _rev(yesterday_start, today)
        weekly_rev   = _rev(week_start, now)
        monthly_rev  = _rev(month_start, now)
        yearly_rev   = _rev(year_start, now)
        total_rev    = float(s.query(sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0.0))
                             .filter(Order.status == OrderStatus.COMPLETED.value).scalar() or 0)

        total_orders = _orders(year_start, now)
        today_orders = _orders(today, now)
        total_users  = s.query(sqlfunc.count(User.id)).scalar() or 1

        avg_order_value = (monthly_rev / max(1, _orders(month_start, now)))
        avg_customer_value = (total_rev / max(1, total_users))

        # Net / Gross Profit: use total revenue - refunds as gross; no cost data = gross = revenue
        refund_amt = float(s.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.amount), 0.0))
                           .filter(Transaction.status == "refunded").scalar() or 0)
        gross_profit = total_rev - refund_amt
        net_profit   = gross_profit  # No cost-of-goods in system; use as gross approximation

    return {
        "today":              today_rev,
        "yesterday":          yest_rev,
        "weekly":             weekly_rev,
        "monthly":            monthly_rev,
        "yearly":             yearly_rev,
        "total":              total_rev,
        "gross_profit":       gross_profit,
        "net_profit":         net_profit,
        "avg_order_value":    avg_order_value,
        "avg_customer_value": avg_customer_value,
        "today_orders":       today_orders,
        "total_orders":       total_orders,
    }


# ─── Sales Forecast (SMA-based) ────────────────────────────────────────────────

def _get_daily_revenues(days: int = 30) -> List[float]:
    """Return last N days of daily revenue figures (oldest first)."""
    today = _utc_today()
    out   = []
    with get_db_session() as s:
        for i in range(days, 0, -1):
            start = today - timedelta(days=i)
            end   = start + timedelta(days=1)
            rev = float(s.query(sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0.0))
                        .filter(Order.status == OrderStatus.COMPLETED.value)
                        .filter(Order.created_at >= start, Order.created_at < end)
                        .scalar() or 0)
            out.append(rev)
    return out


def _sma(series: List[float], window: int) -> float:
    if not series:
        return 0.0
    window_data = series[-window:]
    return sum(window_data) / len(window_data)


def get_sales_forecast() -> Dict[str, Any]:
    """Predict expected daily / weekly / monthly sales using SMA."""
    daily_revs = _get_daily_revenues(30)
    if not daily_revs:
        return {
            "expected_daily": 0, "expected_weekly": 0, "expected_monthly": 0,
            "expected_revenue_30d": 0, "expected_orders_30d": 0,
            "expected_growth_pct": None, "trend": "flat",
            "low_sales_warning": False, "high_sales_trend": False,
        }

    daily_forecast   = _sma(daily_revs, 7)      # 7-day SMA for next day
    weekly_forecast  = daily_forecast * 7
    monthly_forecast = daily_forecast * 30

    # Growth: compare last 7d avg vs prior 7d avg
    last7  = daily_revs[-7:]  if len(daily_revs) >= 7  else daily_revs
    prior7 = daily_revs[-14:-7] if len(daily_revs) >= 14 else daily_revs[:7]
    last7_avg  = sum(last7)  / max(1, len(last7))
    prior7_avg = sum(prior7) / max(1, len(prior7))

    if prior7_avg > 0:
        growth_pct = ((last7_avg - prior7_avg) / prior7_avg) * 100
    else:
        growth_pct = 0.0

    trend = "up" if growth_pct > 5 else ("down" if growth_pct < -5 else "flat")

    # Estimate orders: use same growth ratio applied to historical order count
    with get_db_session() as s:
        today = _utc_today()
        past30_orders = (s.query(sqlfunc.count(Order.id))
                         .filter(Order.status == OrderStatus.COMPLETED.value)
                         .filter(Order.created_at >= today - timedelta(days=30))
                         .scalar() or 0)
    daily_orders_avg = past30_orders / 30.0
    exp_orders_30d   = int(daily_orders_avg * 30 * (1 + growth_pct / 100))

    # Avg revenue of past 30 days for comparison
    all_time_avg = sum(daily_revs) / max(1, len(daily_revs))
    low_warning  = daily_forecast < (all_time_avg * 0.6)
    high_trend   = growth_pct > 20

    return {
        "expected_daily":      round(daily_forecast, 2),
        "expected_weekly":     round(weekly_forecast, 2),
        "expected_monthly":    round(monthly_forecast, 2),
        "expected_revenue_30d": round(monthly_forecast, 2),
        "expected_orders_30d": exp_orders_30d,
        "expected_growth_pct": round(growth_pct, 2),
        "trend":               trend,
        "low_sales_warning":   low_warning,
        "high_sales_trend":    high_trend,
        "confidence_pct":      min(95, max(50, 70 + len(daily_revs) * 0.5)),
    }


def save_forecast_snapshot(period: str = "day") -> None:
    """Generate and persist a forecast snapshot for audit / history."""
    try:
        fc = get_sales_forecast()
        today = _utc_today()
        with get_db_session() as s:
            existing = (s.query(ForecastSnapshot)
                        .filter(ForecastSnapshot.period == period,
                                ForecastSnapshot.forecast_date == today)
                        .first())
            if not existing:
                s.add(ForecastSnapshot(
                    period=period,
                    forecast_date=today,
                    predicted_revenue=fc["expected_daily"] if period == "day"
                        else fc["expected_weekly"] if period == "week"
                        else fc["expected_monthly"],
                    predicted_orders=fc["expected_orders_30d"],
                    predicted_growth_pct=fc["expected_growth_pct"],
                    trend_direction=fc["trend"],
                    confidence_pct=fc["confidence_pct"],
                    model_version="v1_sma",
                ))
                s.commit()
    except Exception:
        logger.exception("save_forecast_snapshot failed")


# ─── Business Insights ────────────────────────────────────────────────────────

def get_best_selling_products(limit: int = 5) -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    OrderItem.product_id,
                    sqlfunc.coalesce(sqlfunc.sum(OrderItem.quantity), 0).label("qty"),
                    sqlfunc.coalesce(sqlfunc.sum(OrderItem.total_price), 0).label("rev"),
                )
                .join(Order, Order.id == OrderItem.order_id)
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .group_by(OrderItem.product_id)
                .order_by(desc("qty"))
                .limit(limit).all())
        result = []
        for pid, qty, rev in rows:
            p = s.query(Product).filter_by(id=pid).first()
            result.append({
                "product_id": pid,
                "name": p.name if p else f"#{pid}",
                "quantity_sold": int(qty),
                "revenue": float(rev),
            })
        return result


def get_worst_selling_products(limit: int = 5) -> List[Dict[str, Any]]:
    with get_db_session() as s:
        sold_ids = {r[0] for r in s.query(OrderItem.product_id)
                    .join(Order, Order.id == OrderItem.order_id)
                    .filter(Order.status == OrderStatus.COMPLETED.value).all()}
        # Products with no sales
        no_sales = (s.query(Product)
                    .filter(Product.is_active == True,  # noqa
                            ~Product.id.in_(sold_ids) if sold_ids else True)
                    .limit(limit).all())
        result = [{"product_id": p.id, "name": p.name, "quantity_sold": 0, "revenue": 0.0}
                  for p in no_sales]

        if len(result) < limit:
            remaining = limit - len(result)
            rows = (s.query(
                        OrderItem.product_id,
                        sqlfunc.coalesce(sqlfunc.sum(OrderItem.quantity), 0).label("qty"),
                        sqlfunc.coalesce(sqlfunc.sum(OrderItem.total_price), 0).label("rev"),
                    )
                    .join(Order, Order.id == OrderItem.order_id)
                    .filter(Order.status == OrderStatus.COMPLETED.value)
                    .group_by(OrderItem.product_id)
                    .order_by(asc("qty"))
                    .limit(remaining).all())
            for pid, qty, rev in rows:
                p = s.query(Product).filter_by(id=pid).first()
                result.append({
                    "product_id": pid,
                    "name": p.name if p else f"#{pid}",
                    "quantity_sold": int(qty),
                    "revenue": float(rev),
                })
        return result


def get_most_active_customers(limit: int = 5) -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    Order.user_id,
                    sqlfunc.count(Order.id).label("order_count"),
                    sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0).label("spend"),
                )
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .group_by(Order.user_id)
                .order_by(desc("order_count"))
                .limit(limit).all())
        result = []
        for uid, cnt, spend in rows:
            u = s.query(User).filter_by(id=uid).first()
            result.append({
                "user_id": uid,
                "telegram_id": u.telegram_id if u else None,
                "username": f"@{u.username}" if u and u.username else f"ID:{uid}",
                "order_count": int(cnt),
                "total_spend": float(spend),
            })
        return result


def get_top_spending_users(limit: int = 5) -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    Order.user_id,
                    sqlfunc.count(Order.id).label("order_count"),
                    sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0).label("spend"),
                )
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .group_by(Order.user_id)
                .order_by(desc("spend"))
                .limit(limit).all())
        result = []
        for uid, cnt, spend in rows:
            u = s.query(User).filter_by(id=uid).first()
            result.append({
                "user_id": uid,
                "telegram_id": u.telegram_id if u else None,
                "username": f"@{u.username}" if u and u.username else f"ID:{uid}",
                "order_count": int(cnt),
                "total_spend": float(spend),
            })
        return result


def get_top_referral_users(limit: int = 5) -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    ReferralReward.referrer_id,
                    sqlfunc.count(ReferralReward.id).label("ref_count"),
                    sqlfunc.coalesce(sqlfunc.sum(ReferralReward.amount), 0).label("earned"),
                )
                .group_by(ReferralReward.referrer_id)
                .order_by(desc("ref_count"))
                .limit(limit).all())
        result = []
        for uid, cnt, earned in rows:
            u = s.query(User).filter_by(id=uid).first()
            result.append({
                "user_id": uid,
                "telegram_id": u.telegram_id if u else None,
                "username": f"@{u.username}" if u and u.username else f"ID:{uid}",
                "referral_count": int(cnt),
                "total_earned": float(earned),
            })
        return result


def get_payment_method_stats() -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    Order.payment_method,
                    sqlfunc.count(Order.id).label("count"),
                    sqlfunc.coalesce(sqlfunc.sum(Order.total_price), 0).label("rev"),
                )
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .group_by(Order.payment_method)
                .order_by(desc("count"))
                .all())
        total = sum(int(r[1]) for r in rows) or 1
        return [
            {
                "method": str(r[0]),
                "count": int(r[1]),
                "revenue": float(r[2]),
                "pct": round(int(r[1]) / total * 100, 1),
            }
            for r in rows
        ]


def get_category_stats() -> List[Dict[str, Any]]:
    with get_db_session() as s:
        rows = (s.query(
                    Product.category_id,
                    sqlfunc.count(OrderItem.id).label("item_count"),
                    sqlfunc.coalesce(sqlfunc.sum(OrderItem.total_price), 0).label("rev"),
                )
                .join(OrderItem, OrderItem.product_id == Product.id)
                .join(Order, Order.id == OrderItem.order_id)
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .group_by(Product.category_id)
                .order_by(desc("rev"))
                .limit(10).all())
        result = []
        for cat_id, cnt, rev in rows:
            cat = s.query(Category).filter_by(id=cat_id).first()
            result.append({
                "category_id": cat_id,
                "name": cat.name if cat else f"#{cat_id}",
                "order_items": int(cnt),
                "revenue": float(rev),
            })
        return result


def get_fastest_growing_product(days: int = 7) -> Optional[Dict[str, Any]]:
    today = _utc_today()
    this_week  = today - timedelta(days=days)
    prior_week = today - timedelta(days=days * 2)
    with get_db_session() as s:
        def _qty(start, end):
            return dict(s.query(
                    OrderItem.product_id,
                    sqlfunc.sum(OrderItem.quantity),
                )
                .join(Order, Order.id == OrderItem.order_id)
                .filter(Order.status == OrderStatus.COMPLETED.value)
                .filter(Order.created_at >= start, Order.created_at < end)
                .group_by(OrderItem.product_id).all())

        this_q  = _qty(this_week, today)
        prior_q = _qty(prior_week, this_week)

        best_pid, best_growth = None, -999
        for pid, q_now in this_q.items():
            q_before = float(prior_q.get(pid, 0) or 0)
            if q_before > 0:
                growth = ((float(q_now) - q_before) / q_before) * 100
            else:
                growth = float(q_now) * 100  # new sales = 100% * units
            if growth > best_growth:
                best_growth, best_pid = growth, pid

        if not best_pid:
            return None
        p = s.query(Product).filter_by(id=best_pid).first()
        return {
            "product_id": best_pid,
            "name": p.name if p else f"#{best_pid}",
            "growth_pct": round(best_growth, 1),
            "units_this_period": int(this_q.get(best_pid, 0)),
        }


# ─── Product Insights ─────────────────────────────────────────────────────────

def get_product_insights() -> Dict[str, List[Dict[str, Any]]]:
    """Return categorized product health insights."""
    with get_db_session() as s:
        all_products = s.query(Product).filter(Product.is_active == True).all()  # noqa
        today = _utc_today()
        week_ago = today - timedelta(days=7)

        # Sales in last 7 days per product
        week_sales = dict(s.query(
                OrderItem.product_id,
                sqlfunc.coalesce(sqlfunc.sum(OrderItem.quantity), 0),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .filter(Order.status == OrderStatus.COMPLETED.value)
            .filter(Order.created_at >= week_ago)
            .group_by(OrderItem.product_id).all())

        # Sales ever per product
        all_sales = dict(s.query(
                OrderItem.product_id,
                sqlfunc.coalesce(sqlfunc.sum(OrderItem.quantity), 0),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .filter(Order.status == OrderStatus.COMPLETED.value)
            .group_by(OrderItem.product_id).all())

        out_of_stock, low_stock, slow_selling, fast_selling, trending, no_sales = [], [], [], [], [], []

        for p in all_products:
            qty    = int(week_sales.get(p.id, 0))
            all_q  = int(all_sales.get(p.id, 0))
            stock  = getattr(p, 'stock_quantity', None)

            if stock is not None:
                if stock == 0:
                    out_of_stock.append({"id": p.id, "name": p.name, "stock": 0})
                elif stock <= 5:
                    low_stock.append({"id": p.id, "name": p.name, "stock": stock})

            if all_q == 0:
                no_sales.append({"id": p.id, "name": p.name})
            elif qty == 0:
                slow_selling.append({"id": p.id, "name": p.name, "week_sales": 0})
            elif qty >= 5:
                fast_selling.append({"id": p.id, "name": p.name, "week_sales": qty})
            if qty >= 10:
                trending.append({"id": p.id, "name": p.name, "week_sales": qty})

        return {
            "out_of_stock":  out_of_stock[:10],
            "low_stock":     low_stock[:10],
            "slow_selling":  slow_selling[:10],
            "fast_selling":  fast_selling[:10],
            "trending":      trending[:10],
            "no_sales":      no_sales[:10],
        }


# ─── Report Generation ────────────────────────────────────────────────────────

def _build_report_data(report_type: str, start: datetime, end: datetime) -> Dict[str, Any]:
    """Collect data for a report. Returns a dict ready for export."""
    data: Dict[str, Any] = {
        "report_type": report_type,
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "generated_at": datetime.utcnow().isoformat(),
    }

    rev = _revenue_in_range(start, end)
    orders = _orders_in_range(start, end)
    data["revenue"] = rev
    data["orders"] = orders
    data["avg_order_value"] = round(rev / max(1, orders), 2)

    if report_type in ("daily", "weekly", "monthly", "yearly", "revenue"):
        data["best_products"] = get_best_selling_products(5)
        data["payment_methods"] = get_payment_method_stats()

    if report_type in ("customer", "weekly", "monthly", "yearly"):
        data["top_customers"] = get_top_spending_users(5)
        data["active_customers"] = get_most_active_customers(5)

    if report_type == "referral":
        data["top_referrers"] = get_top_referral_users(10)

    if report_type == "payment":
        data["payment_breakdown"] = get_payment_method_stats()

    if report_type == "orders":
        with get_db_session() as s:
            rows = (s.query(Order)
                    .filter(Order.created_at >= start, Order.created_at < end)
                    .order_by(desc(Order.created_at)).limit(50).all())
            data["orders_list"] = [
                {"id": o.id, "status": o.status,
                 "total": float(o.total_price or 0),
                 "payment": str(o.payment_method or ""),
                 "created": o.created_at.isoformat() if o.created_at else ""}
                for o in rows
            ]
    return data


def generate_report(report_type: str, admin_tg_id: Optional[int] = None,
                    period_override: Optional[Tuple[datetime, datetime]] = None) -> Dict[str, Any]:
    """Generate a named report and persist it to business_reports."""
    today = _utc_today()
    ranges = {
        "daily":    (today - timedelta(days=1), today),
        "weekly":   (today - timedelta(days=7), today),
        "monthly":  (today - timedelta(days=30), today),
        "yearly":   (today - timedelta(days=365), today),
        "revenue":  (today - timedelta(days=30), today),
        "orders":   (today - timedelta(days=30), today),
        "customer": (today - timedelta(days=30), today),
        "referral": (today - timedelta(days=30), today),
        "payment":  (today - timedelta(days=30), today),
    }
    start, end = period_override or ranges.get(report_type, (today - timedelta(days=30), today))
    data = _build_report_data(report_type, start, end)
    title = f"{report_type.title()} Report {start.strftime('%Y-%m-%d')} – {end.strftime('%Y-%m-%d')}"

    try:
        with get_db_session() as s:
            rpt = BusinessReport(
                report_type=report_type,
                period_start=start,
                period_end=end,
                title=title,
                summary_json=json.dumps({
                    "revenue": data.get("revenue", 0),
                    "orders": data.get("orders", 0),
                    "avg_order_value": data.get("avg_order_value", 0),
                }),
                generated_by=admin_tg_id,
            )
            s.add(rpt)
            s.commit()
    except Exception:
        logger.exception("Failed to save report record")

    data["title"] = title
    return data


# ─── Export functions ─────────────────────────────────────────────────────────

def export_csv(data: Dict[str, Any]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Report", data.get("title", "Business Report")])
    w.writerow(["Generated", data.get("generated_at", "")])
    w.writerow([])
    w.writerow(["Metric", "Value"])
    for k, v in data.items():
        if isinstance(v, (int, float, str)) and k not in ("report_type",):
            w.writerow([k.replace("_", " ").title(), v])
    w.writerow([])
    # Flat list tables
    for key in ("best_products", "top_customers", "active_customers",
                "payment_methods", "top_referrers", "payment_breakdown"):
        rows = data.get(key, [])
        if rows:
            w.writerow([key.replace("_", " ").title()])
            headers = list(rows[0].keys()) if rows else []
            w.writerow(headers)
            for row in rows:
                w.writerow(list(row.values()))
            w.writerow([])
    return buf.getvalue().encode("utf-8-sig")


def export_json(data: Dict[str, Any]) -> bytes:
    return json.dumps(data, indent=2, default=str).encode("utf-8")


def export_excel(data: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    # Title rows
    ws.append(["Report", data.get("title", "Business Report")])
    ws.append(["Generated", data.get("generated_at", "")])
    ws.append([])
    ws.append(["Metric", "Value"])
    ws["A4"].font = header_font
    ws["A4"].fill = header_fill
    ws["B4"].font = header_font
    ws["B4"].fill = header_fill

    row = 5
    for k, v in data.items():
        if isinstance(v, (int, float, str)) and k not in ("report_type",):
            ws.append([k.replace("_", " ").title(), v])
            row += 1

    for key in ("best_products", "top_customers", "active_customers",
                "payment_methods", "top_referrers", "orders_list"):
        rows = data.get(key, [])
        if rows:
            ws.append([])
            ws.append([key.replace("_", " ").title()])
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append(list(r.values()))

    # Auto column widths
    for col in ws.columns:
        max_w = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(data: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    h2_style    = ParagraphStyle("h2",    parent=styles["Heading2"], fontSize=12, spaceAfter=4)
    normal      = styles["Normal"]

    story = [
        Paragraph(data.get("title", "Business Report"), title_style),
        Paragraph(f"Generated: {data.get('generated_at', '')}", normal),
        Spacer(1, 0.4*cm),
    ]

    # Key metrics table
    story.append(Paragraph("Key Metrics", h2_style))
    metrics = [(k.replace("_", " ").title(), str(v))
               for k, v in data.items()
               if isinstance(v, (int, float)) and k not in ("report_type",)]
    if metrics:
        tbl = Table([["Metric", "Value"]] + metrics,
                    colWidths=[9*cm, 7*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF0F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))

    # Sub-tables
    for key, label in [
        ("best_products",   "Best Selling Products"),
        ("top_customers",   "Top Spending Customers"),
        ("payment_methods", "Payment Methods"),
        ("top_referrers",   "Top Referrers"),
    ]:
        rows = data.get(key, [])
        if not rows:
            continue
        story.append(Paragraph(label, h2_style))
        headers = list(rows[0].keys())
        tdata   = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]
        col_w   = [16 / len(headers) * cm] * len(headers)
        tbl2    = Table(tdata, colWidths=col_w)
        tbl2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E74B5")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF0F8")]),
        ]))
        story.append(tbl2)
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    return buf.getvalue()


# ─── Scheduler job ────────────────────────────────────────────────────────────

async def daily_report_job(context) -> None:
    """Telegram job-queue: generate and optionally send daily report."""
    try:
        from utils.bot_config import cfg
        if not cfg.get_bool("biz_auto_daily_report", False):
            return
        save_forecast_snapshot("day")
        generate_report("daily")
        logger.info("Daily business report generated automatically.")
    except Exception:
        logger.exception("daily_report_job failed")


async def weekly_report_job(context) -> None:
    """Telegram job-queue: generate weekly report (run on Mondays)."""
    try:
        from utils.bot_config import cfg
        if not cfg.get_bool("biz_auto_weekly_report", False):
            return
        generate_report("weekly")
        logger.info("Weekly business report generated automatically.")
    except Exception:
        logger.exception("weekly_report_job failed")
