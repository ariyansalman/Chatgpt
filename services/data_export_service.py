"""V43 — Data Export Center service.

Provides centralised, background-capable export of every major data type
in the bot database.  All writes are best-effort: service functions must
never raise exceptions to callers.

Supported export types:
    users, orders, products, categories, transactions, deposits,
    withdrawals, payments, coupons, referrals, broadcasts, flash_sales,
    subscriptions, vip_users, customer_notes, support_tickets,
    activity_timeline, system_logs, statistics, analytics

Supported formats: csv, xlsx, pdf, json, txt, zip
Job lifecycle: pending → running → done | failed
Scheduled jobs: scheduled → running → done | failed
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import threading
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ─── Storage ─────────────────────────────────────────────────────────────────
EXPORT_DIR = Path("/tmp/bot_exports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# ─── Export-type registry ─────────────────────────────────────────────────────
EXPORT_TYPES: dict[str, dict] = {
    "users":            {"label": "👥 Users",              "emoji": "👥"},
    "orders":           {"label": "🧾 Orders",             "emoji": "🧾"},
    "products":         {"label": "📦 Products",           "emoji": "📦"},
    "categories":       {"label": "📂 Categories",         "emoji": "📂"},
    "transactions":     {"label": "💳 Transactions",       "emoji": "💳"},
    "deposits":         {"label": "⬇️ Deposits",            "emoji": "⬇️"},
    "withdrawals":      {"label": "⬆️ Withdrawals",         "emoji": "⬆️"},
    "payments":         {"label": "💰 Payments",           "emoji": "💰"},
    "coupons":          {"label": "🎟 Coupons",            "emoji": "🎟"},
    "referrals":        {"label": "👥 Referral Data",      "emoji": "👥"},
    "broadcasts":       {"label": "📢 Broadcasts",         "emoji": "📢"},
    "flash_sales":      {"label": "⚡ Flash Sales",        "emoji": "⚡"},
    "subscriptions":    {"label": "🔄 Subscriptions",      "emoji": "🔄"},
    "vip_users":        {"label": "👑 VIP Users",          "emoji": "👑"},
    "customer_notes":   {"label": "📝 Customer Notes",     "emoji": "📝"},
    "support_tickets":  {"label": "🎫 Support Tickets",    "emoji": "🎫"},
    "activity_timeline":{"label": "📜 Activity Timeline",  "emoji": "📜"},
    "system_logs":      {"label": "🖥 System Logs",        "emoji": "🖥"},
    "statistics":       {"label": "📊 Statistics",         "emoji": "📊"},
    "analytics":        {"label": "📈 Analytics",          "emoji": "📈"},
}

EXPORT_FORMATS: dict[str, str] = {
    "csv":  "📄 CSV",
    "xlsx": "📊 Excel (.xlsx)",
    "pdf":  "📋 PDF",
    "json": "🗂 JSON",
    "txt":  "📃 TXT",
    "zip":  "🗜 ZIP Archive",
}


# ─── Data fetchers ────────────────────────────────────────────────────────────

def _parse_filters(filters_json: Optional[str]) -> dict:
    if not filters_json:
        return {}
    try:
        return json.loads(filters_json)
    except Exception:
        return {}


def _apply_date_filter(query, model_class, filters: dict):
    """Apply date_from / date_to filter to a SQLAlchemy query."""
    date_col = getattr(model_class, "created_at", None)
    if date_col is None:
        return query
    if filters.get("date_from"):
        try:
            query = query.filter(date_col >= datetime.fromisoformat(filters["date_from"]))
        except Exception:
            pass
    if filters.get("date_to"):
        try:
            query = query.filter(date_col <= datetime.fromisoformat(filters["date_to"]))
        except Exception:
            pass
    return query


def _fetch_users(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import User
    headers = ["id", "telegram_id", "username", "first_name", "last_name",
               "balance", "is_banned", "is_admin", "referral_code", "created_at"]
    q = session.query(User)
    q = _apply_date_filter(q, User, filters)
    if filters.get("status"):
        if filters["status"] == "banned":
            q = q.filter(User.is_banned == True)
        elif filters["status"] == "active":
            q = q.filter(User.is_banned == False)
    rows = []
    for u in q.order_by(User.created_at.desc()).all():
        rows.append([
            u.id, u.telegram_id, u.username or "", u.first_name or "", u.last_name or "",
            float(u.balance or 0), u.is_banned, getattr(u, "is_admin", False),
            getattr(u, "referral_code", "") or "", str(u.created_at or ""),
        ])
    return headers, rows


def _fetch_orders(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Order
    headers = ["id", "user_id", "status", "total_amount", "payment_method",
               "currency", "created_at"]
    q = session.query(Order)
    q = _apply_date_filter(q, Order, filters)
    if filters.get("status"):
        q = q.filter(Order.status == filters["status"])
    rows = []
    for o in q.order_by(Order.created_at.desc()).all():
        rows.append([
            o.id, o.user_id, str(o.status.value if hasattr(o.status, "value") else o.status),
            float(o.total_amount or 0),
            str(o.payment_method.value if hasattr(o.payment_method, "value") else o.payment_method or ""),
            str(getattr(o, "currency", "") or ""), str(o.created_at or ""),
        ])
    return headers, rows


def _fetch_products(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Product
    headers = ["id", "name", "category_id", "price", "stock_quantity",
               "is_active", "product_type", "created_at"]
    q = session.query(Product)
    if filters.get("status"):
        q = q.filter(Product.is_active == (filters["status"] == "active"))
    q = _apply_date_filter(q, Product, filters)
    rows = []
    for p in q.order_by(Product.created_at.desc()).all():
        rows.append([
            p.id, p.name, p.category_id, float(p.price or 0),
            getattr(p, "stock_quantity", 0) or 0, p.is_active,
            str(p.product_type.value if hasattr(p.product_type, "value") else p.product_type or ""),
            str(p.created_at or ""),
        ])
    return headers, rows


def _fetch_categories(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Category
    headers = ["id", "name", "description", "is_active", "created_at"]
    q = session.query(Category)
    rows = []
    for c in q.order_by(Category.id).all():
        rows.append([
            c.id, c.name, getattr(c, "description", "") or "",
            getattr(c, "is_active", True), str(getattr(c, "created_at", "") or ""),
        ])
    return headers, rows


def _fetch_transactions(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Transaction
    headers = ["id", "user_id", "type", "amount", "currency", "status",
               "payment_method", "description", "created_at"]
    q = session.query(Transaction)
    q = _apply_date_filter(q, Transaction, filters)
    if filters.get("status"):
        q = q.filter(Transaction.status == filters["status"])
    rows = []
    for t in q.order_by(Transaction.created_at.desc()).all():
        rows.append([
            t.id, t.user_id,
            str(t.type.value if hasattr(t.type, "value") else t.type or ""),
            float(t.amount or 0),
            str(t.currency.value if hasattr(t.currency, "value") else t.currency or ""),
            str(t.status.value if hasattr(t.status, "value") else t.status or ""),
            str(t.payment_method.value if hasattr(t.payment_method, "value") else t.payment_method or ""),
            getattr(t, "description", "") or "",
            str(t.created_at or ""),
        ])
    return headers, rows


def _fetch_deposits(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import WalletLedger
    headers = ["id", "user_id", "amount", "currency", "source", "reference", "created_at"]
    q = session.query(WalletLedger).filter(WalletLedger.delta > 0)
    q = _apply_date_filter(q, WalletLedger, filters)
    rows = []
    for w in q.order_by(WalletLedger.created_at.desc()).all():
        rows.append([
            w.id, w.user_id, float(w.delta or 0),
            getattr(w, "currency", "") or "",
            getattr(w, "source", "") or "",
            getattr(w, "reference", "") or "",
            str(w.created_at or ""),
        ])
    return headers, rows


def _fetch_withdrawals(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import WalletLedger
    headers = ["id", "user_id", "amount", "currency", "source", "reference", "created_at"]
    q = session.query(WalletLedger).filter(WalletLedger.delta < 0)
    q = _apply_date_filter(q, WalletLedger, filters)
    rows = []
    for w in q.order_by(WalletLedger.created_at.desc()).all():
        rows.append([
            w.id, w.user_id, float(w.delta or 0),
            getattr(w, "currency", "") or "",
            getattr(w, "source", "") or "",
            getattr(w, "reference", "") or "",
            str(w.created_at or ""),
        ])
    return headers, rows


def _fetch_payments(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Transaction
    headers = ["id", "user_id", "amount", "currency", "payment_method",
               "status", "order_id", "created_at"]
    q = session.query(Transaction)
    q = _apply_date_filter(q, Transaction, filters)
    if filters.get("payment_method"):
        q = q.filter(Transaction.payment_method == filters["payment_method"])
    rows = []
    for t in q.order_by(Transaction.created_at.desc()).all():
        rows.append([
            t.id, t.user_id, float(t.amount or 0),
            str(t.currency.value if hasattr(t.currency, "value") else t.currency or ""),
            str(t.payment_method.value if hasattr(t.payment_method, "value") else t.payment_method or ""),
            str(t.status.value if hasattr(t.status, "value") else t.status or ""),
            getattr(t, "order_id", "") or "",
            str(t.created_at or ""),
        ])
    return headers, rows


def _fetch_coupons(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Coupon
    headers = ["id", "code", "discount_type", "discount_value", "max_uses",
               "times_used", "is_active", "expires_at", "created_at"]
    q = session.query(Coupon)
    q = _apply_date_filter(q, Coupon, filters)
    if filters.get("status"):
        q = q.filter(Coupon.is_active == (filters["status"] == "active"))
    rows = []
    for c in q.order_by(Coupon.created_at.desc()).all():
        rows.append([
            c.id, c.code,
            str(c.discount_type.value if hasattr(c.discount_type, "value") else c.discount_type or ""),
            float(c.discount_value or 0),
            getattr(c, "max_uses", 0) or 0,
            getattr(c, "times_used", 0) or 0,
            getattr(c, "is_active", True),
            str(getattr(c, "expires_at", "") or ""),
            str(c.created_at or ""),
        ])
    return headers, rows


def _fetch_referrals(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import ReferralReward
    headers = ["id", "referrer_user_id", "referred_user_id", "reward_amount",
               "is_paid", "created_at"]
    q = session.query(ReferralReward)
    q = _apply_date_filter(q, ReferralReward, filters)
    rows = []
    for r in q.order_by(ReferralReward.created_at.desc()).all():
        rows.append([
            r.id, r.referrer_user_id, r.referred_user_id,
            float(getattr(r, "reward_amount", 0) or 0),
            getattr(r, "is_paid", False),
            str(r.created_at or ""),
        ])
    return headers, rows


def _fetch_broadcasts(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import Broadcast
    headers = ["id", "admin_telegram_id", "message_preview", "sent_count",
               "failed_count", "created_at"]
    q = session.query(Broadcast)
    q = _apply_date_filter(q, Broadcast, filters)
    rows = []
    for b in q.order_by(Broadcast.created_at.desc()).all():
        msg = getattr(b, "message", "") or ""
        rows.append([
            b.id, getattr(b, "admin_telegram_id", "") or "",
            msg[:80], getattr(b, "sent_count", 0) or 0,
            getattr(b, "failed_count", 0) or 0,
            str(b.created_at or ""),
        ])
    return headers, rows


def _fetch_flash_sales(session, filters: dict) -> tuple[list[str], list[list]]:
    try:
        from database.models import FlashSaleEvent
        headers = ["id", "name", "status", "discount_pct", "starts_at", "ends_at", "created_at"]
        q = session.query(FlashSaleEvent)
        q = _apply_date_filter(q, FlashSaleEvent, filters)
        rows = []
        for f in q.order_by(FlashSaleEvent.created_at.desc()).all():
            rows.append([
                f.id, getattr(f, "name", "") or "",
                str(getattr(f, "status", "") or ""),
                float(getattr(f, "discount_pct", 0) or 0),
                str(getattr(f, "starts_at", "") or ""),
                str(getattr(f, "ends_at", "") or ""),
                str(f.created_at or ""),
            ])
        return headers, rows
    except Exception:
        return ["id", "note"], [["N/A", "Flash sales data unavailable"]]


def _fetch_subscriptions(session, filters: dict) -> tuple[list[str], list[list]]:
    try:
        from database.models import Subscription
        headers = ["id", "user_id", "plan_id", "status", "starts_at", "ends_at", "created_at"]
        q = session.query(Subscription)
        q = _apply_date_filter(q, Subscription, filters)
        if filters.get("status"):
            q = q.filter(Subscription.status == filters["status"])
        rows = []
        for s in q.order_by(Subscription.created_at.desc()).all():
            rows.append([
                s.id, s.user_id, getattr(s, "plan_id", "") or "",
                str(getattr(s, "status", "") or ""),
                str(getattr(s, "starts_at", "") or ""),
                str(getattr(s, "ends_at", "") or ""),
                str(s.created_at or ""),
            ])
        return headers, rows
    except Exception:
        return ["id", "note"], [["N/A", "Subscriptions data unavailable"]]


def _fetch_vip_users(session, filters: dict) -> tuple[list[str], list[list]]:
    try:
        from database.models import UserVipTier, VipTier, User
        from sqlalchemy.orm import joinedload
        headers = ["user_id", "username", "telegram_id", "tier_name",
                   "points", "assigned_at"]
        q = (session.query(UserVipTier)
             .join(User, UserVipTier.user_id == User.id)
             .join(VipTier, UserVipTier.tier_id == VipTier.id))
        rows = []
        for uvt in q.order_by(UserVipTier.assigned_at.desc()).all():
            u = uvt.user if hasattr(uvt, "user") else None
            tier = uvt.tier if hasattr(uvt, "tier") else None
            rows.append([
                uvt.user_id,
                (u.username or "") if u else "",
                (u.telegram_id or "") if u else "",
                (tier.name or "") if tier else "",
                float(getattr(uvt, "points", 0) or 0),
                str(getattr(uvt, "assigned_at", "") or ""),
            ])
        return headers, rows
    except Exception:
        return ["id", "note"], [["N/A", "VIP user data unavailable"]]


def _fetch_customer_notes(session, filters: dict) -> tuple[list[str], list[list]]:
    try:
        from database.models import CustomerNote
        headers = ["id", "user_id", "admin_telegram_id", "note", "created_at"]
        q = session.query(CustomerNote)
        q = _apply_date_filter(q, CustomerNote, filters)
        rows = []
        for n in q.order_by(CustomerNote.created_at.desc()).all():
            rows.append([
                n.id, n.user_id,
                getattr(n, "admin_telegram_id", "") or "",
                (getattr(n, "note", "") or "")[:200],
                str(n.created_at or ""),
            ])
        return headers, rows
    except Exception:
        return ["id", "note"], [["N/A", "Customer notes unavailable"]]


def _fetch_support_tickets(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import SupportTicket
    headers = ["id", "user_id", "subject", "status", "priority", "created_at", "updated_at"]
    q = session.query(SupportTicket)
    q = _apply_date_filter(q, SupportTicket, filters)
    if filters.get("status"):
        q = q.filter(SupportTicket.status == filters["status"])
    rows = []
    for t in q.order_by(SupportTicket.created_at.desc()).all():
        rows.append([
            t.id, t.user_id, getattr(t, "subject", "") or "",
            str(t.status.value if hasattr(t.status, "value") else t.status or ""),
            str(t.priority.value if hasattr(t.priority, "value") else t.priority or ""),
            str(t.created_at or ""),
            str(getattr(t, "updated_at", "") or ""),
        ])
    return headers, rows


def _fetch_activity_timeline(session, filters: dict) -> tuple[list[str], list[list]]:
    try:
        from database.models import GlobalActivityEntry
        headers = ["id", "user_id", "username", "action", "category",
                   "description", "status", "created_at"]
        q = session.query(GlobalActivityEntry)
        q = _apply_date_filter(q, GlobalActivityEntry, filters)
        if filters.get("category"):
            q = q.filter(GlobalActivityEntry.category == filters["category"])
        rows = []
        for e in q.order_by(GlobalActivityEntry.created_at.desc()).all():
            rows.append([
                e.id, e.user_id or "", e.username or "", e.action,
                e.category, (e.description or "")[:150], e.status,
                str(e.created_at or ""),
            ])
        return headers, rows
    except Exception:
        return ["id", "note"], [["N/A", "Activity timeline unavailable"]]


def _fetch_system_logs(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import AdminAuditLog
    headers = ["id", "admin_telegram_id", "action", "target_user_id",
               "details", "created_at"]
    q = session.query(AdminAuditLog)
    q = _apply_date_filter(q, AdminAuditLog, filters)
    rows = []
    for a in q.order_by(AdminAuditLog.created_at.desc()).all():
        rows.append([
            a.id, a.admin_telegram_id, a.action,
            getattr(a, "target_user_id", "") or "",
            (getattr(a, "details", "") or "")[:200],
            str(a.created_at or ""),
        ])
    return headers, rows


def _fetch_statistics(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import User, Order, Product, Transaction
    headers = ["metric", "value"]
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    rows = [
        ["Total Users",           session.query(User).count()],
        ["Active Users (30d)",    session.query(User).filter(
            User.created_at >= now - timedelta(days=30)).count()],
        ["Total Orders",          session.query(Order).count()],
        ["Orders This Month",     session.query(Order).filter(
            Order.created_at >= month_start).count()],
        ["Total Products",        session.query(Product).count()],
        ["Active Products",       session.query(Product).filter(
            Product.is_active == True).count()],
        ["Total Transactions",    session.query(Transaction).count()],
        ["Report Generated At",   str(now)],
    ]
    return headers, rows


def _fetch_analytics(session, filters: dict) -> tuple[list[str], list[list]]:
    from database.models import User, Order, Transaction
    headers = ["period", "new_users", "orders", "revenue"]
    now = datetime.utcnow()
    rows = []
    for days_ago in range(30, -1, -1):
        day = (now - timedelta(days=days_ago)).date()
        day_start = datetime(day.year, day.month, day.day)
        day_end = day_start + timedelta(days=1)
        new_users = session.query(User).filter(
            User.created_at >= day_start, User.created_at < day_end).count()
        orders = session.query(Order).filter(
            Order.created_at >= day_start, Order.created_at < day_end).count()
        revenue_q = session.query(Transaction).filter(
            Transaction.created_at >= day_start, Transaction.created_at < day_end)
        revenue = sum(float(t.amount or 0) for t in revenue_q.all() if float(t.amount or 0) > 0)
        rows.append([str(day), new_users, orders, round(revenue, 2)])
    return headers, rows


# ─── Dispatch table ───────────────────────────────────────────────────────────
_FETCHERS = {
    "users":             _fetch_users,
    "orders":            _fetch_orders,
    "products":          _fetch_products,
    "categories":        _fetch_categories,
    "transactions":      _fetch_transactions,
    "deposits":          _fetch_deposits,
    "withdrawals":       _fetch_withdrawals,
    "payments":          _fetch_payments,
    "coupons":           _fetch_coupons,
    "referrals":         _fetch_referrals,
    "broadcasts":        _fetch_broadcasts,
    "flash_sales":       _fetch_flash_sales,
    "subscriptions":     _fetch_subscriptions,
    "vip_users":         _fetch_vip_users,
    "customer_notes":    _fetch_customer_notes,
    "support_tickets":   _fetch_support_tickets,
    "activity_timeline": _fetch_activity_timeline,
    "system_logs":       _fetch_system_logs,
    "statistics":        _fetch_statistics,
    "analytics":         _fetch_analytics,
}


# ─── File writers ─────────────────────────────────────────────────────────────

def _write_csv(headers: list, rows: list, path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def _write_xlsx(headers: list, rows: list, path: Path, sheet_name: str = "Export") -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=str(h))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_i, row in enumerate(rows, 2):
        for col_i, val in enumerate(row, 1):
            ws.cell(row=row_i, column=col_i, value=str(val) if val is not None else "")
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    wb.save(path)


def _write_pdf(headers: list, rows: list, path: Path, title: str = "Export Report") -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4),
                             leftMargin=1*cm, rightMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} — {len(rows):,} records",
                  styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    col_count = len(headers)
    # Limit to first 1000 rows to keep PDF manageable
    display_rows = rows[:1000]
    table_data = [[str(h) for h in headers]] + [[str(v)[:80] for v in r] for r in display_rows]
    col_width = max(1.5*cm, (landscape(A4)[0] - 2*cm) / max(col_count, 1))
    col_widths = [col_width] * col_count
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F8")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    if len(rows) > 1000:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(f"Note: PDF limited to first 1,000 of {len(rows):,} records. Use CSV/XLSX for full export.", styles["Normal"]))
    doc.build(story)


def _write_json(headers: list, rows: list, path: Path) -> None:
    data = [dict(zip(headers, [str(v) if v is not None else None for v in row])) for row in rows]
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"exported_at": datetime.utcnow().isoformat(), "count": len(data), "records": data},
                  f, ensure_ascii=False, indent=2, default=str)


def _write_txt(headers: list, rows: list, path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Export Report — {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n")
        f.write(f"Total records: {len(rows):,}\n")
        f.write("=" * 80 + "\n\n")
        col_widths = [max(len(str(h)), max((len(str(r[i])) for r in rows), default=0))
                      for i, h in enumerate(headers)]
        col_widths = [min(w, 30) for w in col_widths]
        header_line = "  ".join(str(h).ljust(col_widths[i]) for i, h in enumerate(headers))
        f.write(header_line + "\n")
        f.write("-" * len(header_line) + "\n")
        for row in rows:
            f.write("  ".join(str(v)[:col_widths[i]].ljust(col_widths[i])
                               for i, v in enumerate(row)) + "\n")


def _write_zip(headers: list, rows: list, path: Path, export_type: str) -> None:
    """Write a ZIP containing CSV + JSON."""
    csv_buf = io.StringIO()
    w = csv.writer(csv_buf)
    w.writerow(headers)
    w.writerows(rows)
    json_data = [dict(zip(headers, [str(v) for v in row])) for row in rows]
    json_str = json.dumps({"exported_at": datetime.utcnow().isoformat(),
                            "count": len(rows), "records": json_data},
                           ensure_ascii=False, indent=2, default=str)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{export_type}.csv", csv_buf.getvalue())
        zf.writestr(f"{export_type}.json", json_str)


# ─── Job management ───────────────────────────────────────────────────────────

def create_job(admin_telegram_id: int, export_type: str, fmt: str,
               filters: Optional[dict] = None,
               scheduled_at: Optional[datetime] = None,
               label: Optional[str] = None) -> Optional[int]:
    """Create an ExportJob record. Returns job_id or None on error."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        from utils.audit import log_admin_action
        status = "scheduled" if scheduled_at else "pending"
        filters_json = json.dumps(filters or {})
        job_label = label or f"{EXPORT_TYPES.get(export_type, {}).get('label', export_type)} ({fmt.upper()})"
        with get_db_session() as session:
            job = ExportJob(
                admin_telegram_id=admin_telegram_id,
                export_type=export_type,
                format=fmt,
                status=status,
                filters=filters_json,
                label=job_label,
                scheduled_at=scheduled_at,
                created_at=datetime.utcnow(),
            )
            session.add(job)
            session.flush()
            job_id = job.id
        try:
            log_admin_action(admin_telegram_id, "dec_create_job",
                             details=f"type={export_type} fmt={fmt} status={status}")
        except Exception:
            pass
        return job_id
    except Exception as e:
        logger.error("create_job error: %s", e, exc_info=True)
        return None


def get_job(job_id: int) -> Optional[dict]:
    """Return a job as a plain dict, or None if not found."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        with get_db_session() as session:
            job = session.query(ExportJob).filter(ExportJob.id == job_id).first()
            if not job:
                return None
            return {
                "id": job.id,
                "admin_telegram_id": job.admin_telegram_id,
                "export_type": job.export_type,
                "format": job.format,
                "status": job.status,
                "filters": _parse_filters(job.filters),
                "file_path": job.file_path,
                "file_size": job.file_size,
                "row_count": job.row_count,
                "error_message": job.error_message,
                "label": job.label,
                "scheduled_at": job.scheduled_at,
                "started_at": job.started_at,
                "completed_at": job.completed_at,
                "created_at": job.created_at,
            }
    except Exception as e:
        logger.error("get_job error: %s", e, exc_info=True)
        return None


def list_jobs(admin_telegram_id: Optional[int] = None,
              status: Optional[str] = None,
              limit: int = 30,
              offset: int = 0) -> list[dict]:
    """Return a list of export jobs as plain dicts."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        with get_db_session() as session:
            q = session.query(ExportJob)
            if admin_telegram_id:
                q = q.filter(ExportJob.admin_telegram_id == admin_telegram_id)
            if status:
                q = q.filter(ExportJob.status == status)
            jobs = q.order_by(ExportJob.created_at.desc()).offset(offset).limit(limit).all()
            return [{
                "id": j.id,
                "export_type": j.export_type,
                "format": j.format,
                "status": j.status,
                "label": j.label,
                "file_size": j.file_size,
                "row_count": j.row_count,
                "error_message": j.error_message,
                "scheduled_at": j.scheduled_at,
                "completed_at": j.completed_at,
                "created_at": j.created_at,
            } for j in jobs]
    except Exception as e:
        logger.error("list_jobs error: %s", e, exc_info=True)
        return []


def count_jobs(admin_telegram_id: Optional[int] = None,
               status: Optional[str] = None) -> int:
    try:
        from database import get_db_session
        from database.models import ExportJob
        with get_db_session() as session:
            q = session.query(ExportJob)
            if admin_telegram_id:
                q = q.filter(ExportJob.admin_telegram_id == admin_telegram_id)
            if status:
                q = q.filter(ExportJob.status == status)
            return q.count()
    except Exception:
        return 0


def delete_job(job_id: int) -> bool:
    """Delete a job record and its file. Returns True on success."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        with get_db_session() as session:
            job = session.query(ExportJob).filter(ExportJob.id == job_id).first()
            if not job:
                return False
            if job.file_path:
                try:
                    Path(job.file_path).unlink(missing_ok=True)
                except Exception:
                    pass
            session.delete(job)
        return True
    except Exception as e:
        logger.error("delete_job error: %s", e, exc_info=True)
        return False


def _set_job_status(job_id: int, **kwargs) -> None:
    try:
        from database import get_db_session
        from database.models import ExportJob
        with get_db_session() as session:
            job = session.query(ExportJob).filter(ExportJob.id == job_id).first()
            if job:
                for k, v in kwargs.items():
                    setattr(job, k, v)
    except Exception as e:
        logger.error("_set_job_status error: %s", e, exc_info=True)


def _run_export(job_id: int) -> None:
    """Execute the export in-process (called inside a thread)."""
    _set_job_status(job_id, status="running", started_at=datetime.utcnow())
    job = get_job(job_id)
    if not job:
        return
    try:
        from database import get_db_session
        fetcher = _FETCHERS.get(job["export_type"])
        if not fetcher:
            _set_job_status(job_id, status="failed",
                            error_message=f"Unknown export type: {job['export_type']}")
            return

        with get_db_session() as session:
            headers, rows = fetcher(session, job["filters"] or {})

        fmt = job["format"]
        ext = "zip" if fmt == "zip" else fmt
        filename = f"export_{job['export_type']}_{job_id}_{int(time.time())}.{ext}"
        path = EXPORT_DIR / filename

        if fmt == "csv":
            _write_csv(headers, rows, path)
        elif fmt == "xlsx":
            _write_xlsx(headers, rows, path, sheet_name=job["export_type"][:31])
        elif fmt == "pdf":
            title = f"{EXPORT_TYPES.get(job['export_type'], {}).get('label', job['export_type'])} Export"
            _write_pdf(headers, rows, path, title=title)
        elif fmt == "json":
            _write_json(headers, rows, path)
        elif fmt == "txt":
            _write_txt(headers, rows, path)
        elif fmt == "zip":
            _write_zip(headers, rows, path, job["export_type"])
        else:
            _set_job_status(job_id, status="failed",
                            error_message=f"Unknown format: {fmt}")
            return

        file_size = path.stat().st_size if path.exists() else 0
        _set_job_status(
            job_id,
            status="done",
            file_path=str(path),
            file_size=file_size,
            row_count=len(rows),
            completed_at=datetime.utcnow(),
        )
        logger.info("Export job %d done: %s (%d rows, %d bytes)", job_id, filename, len(rows), file_size)
    except Exception as e:
        logger.error("Export job %d failed: %s", job_id, e, exc_info=True)
        _set_job_status(job_id, status="failed", error_message=str(e)[:500])


def start_job(job_id: int) -> None:
    """Launch _run_export in a background daemon thread."""
    t = threading.Thread(target=_run_export, args=(job_id,), daemon=True,
                         name=f"dec-export-{job_id}")
    t.start()


async def process_scheduled_jobs(context=None) -> None:
    """Kick off any scheduled jobs whose scheduled_at has passed.
    Called from a periodic job_queue job in bot.py."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        now = datetime.utcnow()
        with get_db_session() as session:
            due = (session.query(ExportJob)
                   .filter(ExportJob.status == "scheduled",
                           ExportJob.scheduled_at <= now)
                   .all())
            ids = [j.id for j in due]
        for jid in ids:
            _set_job_status(jid, status="pending")
            start_job(jid)
    except Exception as e:
        logger.error("process_scheduled_jobs error: %s", e, exc_info=True)


# ─── Stats ────────────────────────────────────────────────────────────────────

def get_stats() -> dict:
    """Return export statistics for the admin panel."""
    try:
        from database import get_db_session
        from database.models import ExportJob
        now = datetime.utcnow()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        month_start = now - timedelta(days=30)
        with get_db_session() as session:
            total = session.query(ExportJob).count()
            today = session.query(ExportJob).filter(ExportJob.created_at >= today_start).count()
            weekly = session.query(ExportJob).filter(ExportJob.created_at >= week_start).count()
            monthly = session.query(ExportJob).filter(ExportJob.created_at >= month_start).count()
            failed = session.query(ExportJob).filter(ExportJob.status == "failed").count()
            pending = session.query(ExportJob).filter(ExportJob.status.in_(["pending", "running", "scheduled"])).count()
            # Largest export by file_size
            largest = (session.query(ExportJob)
                       .filter(ExportJob.file_size != None)
                       .order_by(ExportJob.file_size.desc())
                       .first())
            recent = (session.query(ExportJob)
                      .filter(ExportJob.status == "done")
                      .order_by(ExportJob.completed_at.desc())
                      .first())
        return {
            "total": total,
            "today": today,
            "weekly": weekly,
            "monthly": monthly,
            "failed": failed,
            "pending": pending,
            "largest_size": largest.file_size if largest else 0,
            "largest_type": largest.export_type if largest else "",
            "recent_label": recent.label if recent else "",
            "recent_at": recent.completed_at if recent else None,
        }
    except Exception as e:
        logger.error("get_stats error: %s", e, exc_info=True)
        return {"total": 0, "today": 0, "weekly": 0, "monthly": 0,
                "failed": 0, "pending": 0, "largest_size": 0,
                "largest_type": "", "recent_label": "", "recent_at": None}
